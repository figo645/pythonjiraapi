import requests
import json
import re
import sys
import string
import random
import os
import pytz
import unicodedata
import csv
import openpyxl
import logging
from chardet import detect
sys.path.append("../..")
from datetime import date, timedelta, datetime


# for input

# 保存查询类型 repo 还是 project
def get_query_type():
    query_type = input("请输入查询类型, repo 还是 project 的数据, 如果不输入, 默认查询repo: ")
    while query_type not in ["repo", "project"] and len(query_type) != 0:
        query_type = input("请重新输入查询类型, repo 还是 project 的数据, 如果不输入, 默认查询repo: ")

    if len(query_type) == 0:
        query_type = 'repo'

    print(f"您输入的 查询类型 是 {query_type}")

    return query_type




# 保存是否查询特定邮箱的数据, 返回 True 或者 False
def get_whether_to_query_specific_email():
    whether = input("请输入是否查询特定邮箱的数据, False 还是 True, False选择1, True选择2, 如果不输入, 默认False: ")
    while whether not in ['1', '2'] and len(whether) != 0:
        whether = input("请重新输入是否查询特定邮箱的数据, False 还是 True, False选择1, True选择2, 如果不输入, 默认False: ")

    whether = True if whether == '2' else False

    print(f"您输入的 是否查询特定邮箱的数据 是 {whether}")

    return whether



# 保存用户输入的特定邮箱, 返回特定邮箱的列表
def get_specific_emails():
    specific_emails = input("请输入您要查询的邮箱列表, 间隔符号支持英文 |, 不支持输入为空: ")
    while len(specific_emails) == 0:
        specific_emails = input("请重新输入您要查询的邮箱列表, 间隔符号支持英文 |, 不支持输入为空: ")
    specific_emails_list = specific_emails.split('|')

    return specific_emails_list



# 保存用户输入的offset
def get_offset():
    offset = input("请输入起始查询位置offset，代表想要从第(offset+1)名 (offset仅支持自然数) 开始查询数据，如果输入为空，则从第1位开始查询")
    while not is_natural_number(offset) and len(offset) != 0:
        offset = input("请重新输入起始查询位置offset，代表想要从第(offset+1)名 (offset仅支持自然数) 开始查询行数据，如果输入为空，则从第1位开始查询")

    if len(offset) == 0:
        offset = 0
    else:
        offset = int(offset)
    print(f"您输入的起始查询位置offset 是 {offset}, 代表想要从 第{offset+1}名 开始数据")
    return offset



# 保存用户输入的limit
def get_limit():
    limit = input("请输入查询数据的条数/限制limit，代表想要查询多少条数据（表示取的条数），如果输入为空，则查询1000条数据")
    while not is_positive_integer(limit) and len(limit) != 0:
        limit = input("请重新输入查询数据的条数/限制limit，代表想要查询多少条数据（表示取的条数），如果输入为空，则查询1000条数据")

    if len(limit) == 0:
        limit = 1000
    else:
        limit = int(limit)
    print(f"您输入的查询数据的条数/限制 是 {limit}，代表想要查询多少条数据（表示取的条数）")
    return limit



# 保存用户输入的是否包含元信息
def get_includingMetaData_allow_empty():
    result = input(f"返回中是否包含元信息，可以输入true或者输入false，若不输入则为false: ")
    while len(result) != 0 and not(result in ['true', 'false']):
        result = input(f"请重新输入返回中是否包含元信息，可以输入true或者输入false，若不输入则为false:")

    if result == 'true':
        result = True
    elif result == 'false' or len(result) == 0:
        result = False

    print(f"返回中是否包含元信息 是 {result}")
    return result


# 10_Sprint
# 保存用户输入的迭代ID，或 迭代名称(支持正则匹配，间隔符号仅支持英文|) 或 all(代表查询全部)
def get_sprint_id_or_name_pattern():
    # 以下内容根据需要填写
    special_sprint_regex = input("请输入特定的迭代ID 或 迭代名称(支持正则匹配, 间隔符号仅支持英文|) 或 all(代表查询全部迭代的事务列表), 不支持输入为空: ")
    while len(special_sprint_regex) == 0:
        special_sprint_regex = input("请重新输入特定的迭代ID 或 迭代名称(支持正则匹配, 间隔符号仅支持英文|) 或 all(代表查询全部迭代的事务列表), 不支持输入为空: ")

    sprint_pattern = re.compile(special_sprint_regex, re.IGNORECASE )

    if len(special_sprint_regex) != 'all':
        print(f"您输入的 特定的迭代(组)ID 或 迭代名称 是 {special_sprint_regex}")
    else:
        print(f"您输入的是 'all', 稍后将查询全部迭代的事务列表数据")

    return special_sprint_regex, sprint_pattern



# 使用迭代ID 或 迭代名称 或 all 检索迭代
def find_sprint_by_id_or_name(special_sprint_regex, sprint_pattern, sprint_list_dict):
    # 要查询的迭代的ids
    # sprint_list = get_sprint_list(client)
    # print(f"sprint_list_dict = {sprint_list_dict}")
    # list to dict
    sprint_dict = {}

    if special_sprint_regex == 'all':
        for index, i in enumerate(sprint_list_dict.items()):
            ran_str = ''.join(random.sample(string.ascii_letters + string.digits, 20))
            sprint_dict[ran_str] = {}
            sprint_dict[ran_str] = i[1]
            sprint_dict[ran_str]['迭代ID'] = i[1]['id']
            sprint_dict[ran_str]['检索命中字段'] = "列出全部迭代"
            sprint_dict[ran_str]['被检索字段'] = "列出全部迭代"
            sprint_dict[ran_str]['迭代名称'] = i[1]['name']

    else:
        for index_2, i_2 in enumerate(sprint_list_dict.items()):
            ran_str = ''.join(random.sample(string.ascii_letters + string.digits, 20))
            if sprint_pattern.findall(i_2[1]['id']):
                sprint_dict[ran_str] = {}
                sprint_dict[ran_str] = i_2[1]
                sprint_dict[ran_str]['迭代ID'] = i_2[1]['id']
                sprint_dict[ran_str]['检索命中字段'] = '迭代ID'

            if sprint_pattern.findall(i_2[1]['name']):
                if i_2[1]['id'] in sprint_dict.keys():
                    origin_str = sprint_dict[i_2[1]['id']]['检索命中字段']
                    sprint_dict[ran_str]['检索命中字段'] = f"{origin_str} & 迭代名称"
                else:
                    sprint_dict[ran_str] = {}
                    sprint_dict[ran_str] = i_2[1]
                    sprint_dict[ran_str]['迭代ID'] = i_2[1]['id']
                    sprint_dict[ran_str]['检索命中字段'] = '迭代名称'

            if sprint_pattern.findall(i_2[1]['id']) or sprint_pattern.findall(i_2[1]['name']):
                sprint_dict[ran_str]['被检索字段'] = special_sprint_regex
                sprint_dict[ran_str]['迭代名称'] = i_2[1]['name']

    if len(sprint_dict) == 0:
        print(f"!!!!!!很抱歉, 系统没有找到匹配迭代ID、迭代名称 正则表达式的: {special_sprint_regex} 的迭代, 请您重新执行脚本文件, 并修改迭代信息, 以便找到您需要的迭代")

    print(f"目前累计检索到的迭代数量是: {len(sprint_dict.keys())}")
    # print(f"/n")
    # print(f"sprint_dict = {sprint_dict}")
    return sprint_dict



# 输入排序字段(支持输入: "key" "worktime" "dev_eq")， key-事务号，worktime-工时，dev_eq-代码当量，支持输入为空
# issue_field = get_issue_field()
def get_issue_field():
    issue_field = input("请输入排序字段(支持输入: 'key' 'worktime' 'dev_eq')， key-事务号，worktime-工时，dev_eq-代码当量，支持输入为空: ")
    while issue_field not in ["key", "worktime", "dev_eq"] and len(issue_field) != 0:
        issue_field = input("请重新请输入排序字段(支持输入: 'key' 'worktime' 'dev_eq')， key-事务号，worktime-工时，dev_eq-代码当量，支持输入为空:")

    if len(issue_field) != 0:
        print(f"您输入的排序字段 是 {issue_field}")
    else:
        print(f"您 输入的排序字段为空")
    return issue_field



# 输入排序顺序(支持输入: "ASC" "DESC")，ASC 表示升序排列，即数据按照从小到大的顺序进行排序；而 DESC 表示降序排列，即数据按照从大到小的顺序进行排序，支持输入为空
# issue_sort = get_issue_sort()
def get_issue_sort():
    issue_sort = input("请输入排序顺序(支持输入: 'ASC' 'DESV'), ASC 表示升序排列，即数据按照从小到大的顺序进行排序；而 DESC 表示降序排列，即数据按照从大到小的顺序进行排序，支持输入为空: ")
    while issue_sort not in ["ASC", "DESC"] and len(issue_sort) != 0:
        issue_sort = input("请重新请输入排序顺序(支持输入: 'ASC' 'DESV'), ASC 表示升序排列，即数据按照从小到大的顺序进行排序；而 DESC 表示降序排列，即数据按照从大到小的顺序进行排序，支持输入为空:")

    if len(issue_sort) != 0:
        print(f"您输入的排序顺序 是 {issue_sort}")
    else:
        print(f"您 输入的排序顺序为空")
    return issue_sort



# 输入6：事务优先级priority，从sprint的元信息中获取，支持输入为空，如果输入为空，代表查询全部
# issue_priority = get_issue_priority()
def get_issue_priority():
    issue_priority = input("请输入事务优先级priority，从sprint的元信息中获取，支持输入为空，如果输入为空，代表查询全部: ")

    if len(issue_priority) != 0:
        print(f"您输入事务优先级priority 是 {issue_priority}")
    else:
        print(f"您 输入事务优先级priority为空")
    return issue_priority



# 输入7：事务处理人assigneeId，从sprint的元信息中获取，支持输入为空，如果输入为空，代表查询全部
# issue_assigneeId = get_issue_assigneeId()
def get_issue_assigneeId():
    issue_assigneeId = input("请输入事务处理人assigneeId，从sprint的元信息中获取，支持输入为空，如果输入为空，代表查询全部: ")

    if len(issue_assigneeId) != 0:
        print(f"您输入事务处理人assigneeId 是 {issue_assigneeId}")
    else:
        print(f"您 输入事务处理人assigneeId为空")
    return issue_assigneeId



# 输入8：事务状态名称status，从sprint的元信息中获取，支持输入为空，如果输入为空，代表查询全部
# issue_status = get_issue_status()
def get_issue_status():
    issue_status = input("请输入事务状态名称status，从sprint的元信息中获取，支持输入为空，如果输入为空，代表查询全部: ")

    if len(issue_status) != 0:
        print(f"您输入事务状态名称status 是 {issue_status}")
    else:
        print(f"您 输入事务状态名称status为空")
    return issue_status




# 输入9：事务类型名称type，从sprint的元信息中获取，支持输入为空，如果输入为空，代表查询全部
# issue_type = get_issue_type()
def get_issue_type():
    issue_type = input("请输入事务类型名称type，从sprint的元信息中获取，支持输入为空，如果输入为空，代表查询全部: ")

    if len(issue_type) != 0:
        print(f"您输入事务类型名称type 是 {issue_type}")
    else:
        print(f"您 输入事务类型名称type为空")
    return issue_type




# 输入10：事务唯一识别符列表issueKeys，这里使用精确匹配，如果输入为空，代表查询全部
# issueKeys = get_issueKeys()
def get_issueKeys():
    issueKeys = input("请输入事务唯一识别符列表issueKeys，这里使用精确匹配，如果输入为空，代表查询全部: ")

    if len(issueKeys) != 0:
        print(f"您输入事务唯一识别符列表issueKeys 是 {issueKeys}")
    else:
        print(f"您 事务唯一识别符列表issueKeys为空")
    return issueKeys



# 保存用户输入的csv_type
def get_csv_type():
    csv_type = input("请输入csv导入类型(支持输入: 'issues' 'issue_repo_commits' ), 不支持输入为空: ")
    while csv_type not in ["issues", "issue_repo_commits"] and len(csv_type) != 0:
        csv_type = input("请重新输入csv导入类型(支持输入: 'issues' 'issue_repo_commits' ), 不支持输入为空:")
    if len(csv_type) != 0:
        print(f"您输入的csv导入类型( 是 {csv_type}")
    return csv_type



# 11_IssueTracking
def get_issue_field_with_date():
    issue_field_with_date = input("请输入排序字段field: 'created_date' 'updated_date')， created_date-创建日期，updated_date-更新日期，支持输入为空: ")
    while issue_field_with_date not in ["created_date", "updated_date"] and len(issue_field_with_date) != 0:
        issue_field_with_date = input("请重新请输入排序字段field: 'created_date' 'updated_date')， created_date-创建日期，updated_date-更新日期，支持输入为空:")

    if len(issue_field_with_date) != 0:
        print(f"您输入的排序字段 是 {issue_field_with_date}")
    else:
        print(f"您 输入的排序字段为空")
    return issue_field_with_date



# 输入事务状态的类别statusCategory(支持输入: "TODO" "IN_PROGRESS" "DONE")，TODO 表示处于等待阶段的事务，IN_PROGRESS 表示处在处理中阶段的事务，DONE表示已完成的事务，支持输入为空
def get_statusCategory():
    statusCategory = input("请输入事务状态的类别(支持输入: 'TODO' 'IN_PROGRESS' 'DONE'), 支持输入为空: ")
    while statusCategory not in ["TODO", "IN_PROGRESS", "DONE"] and len(statusCategory) != 0:
        statusCategory = input("请重输入事务状态的类别(支持输入: 'TODO' 'IN_PROGRESS' 'DONE'), 支持输入为空:")
    if len(statusCategory) != 0:
        print(f"您输入的事务状态的类别 是 {statusCategory}")
    else:
        print(f"您输入的事务状态的类别 为空")
    return statusCategory



# 事务类型的类别typeCategory(支持输入: "BUG" "REQUIREMENT" "INCIDENT" "TECH_DEBT")，BUG 表示事务类型是缺陷的事务，REQUIREMENT 表示事务类型是需求的事务，INCIDENT表示事务类型是事故的事务，TECH_DEBT表示事务类型是技术债的事务，支持输入为空
def get_typeCategory():
    typeCategory = input("请输入事务类型的类别(支持输入: 'BUG' 'REQUIREMENT' 'INCIDENT' 'TECH_DEBT'), 支持输入为空: ")
    while typeCategory not in ["TODO", "IN_PROGRESS", "DONE"] and len(typeCategory) != 0:
        typeCategory = input("请重输入事务类型的类别(支持输入: 'BUG' 'REQUIREMENT' 'INCIDENT' 'TECH_DEBT'), 支持输入为空:")
    if len(typeCategory) != 0:
        print(f"您输入的事务类型的类别 是 {typeCategory}")
    else:
        print(f"您输入的事务类型的类别 为空")
    return typeCategory



# 事务类型的类别typeCategory(支持输入: "BUG" "REQUIREMENT" "INCIDENT" "TECH_DEBT")，BUG 表示事务类型是缺陷的事务，REQUIREMENT 表示事务类型是需求的事务，INCIDENT表示事务类型是事故的事务，TECH_DEBT表示事务类型是技术债的事务，支持输入为空
def get_typeCategory():
    typeCategory = input("请输入事务类型的类别(支持输入: 'BUG' 'REQUIREMENT' 'INCIDENT' 'TECH_DEBT'), 支持输入为空: ")
    while typeCategory not in ["TODO", "IN_PROGRESS", "DONE"] and len(typeCategory) != 0:
        typeCategory = input("请重输入事务类型的类别(支持输入: 'BUG' 'REQUIREMENT' 'INCIDENT' 'TECH_DEBT'), 支持输入为空:")
    if len(typeCategory) != 0:
        print(f"您输入的事务类型的类别 是 {typeCategory}")
    else:
        print(f"您输入的事务类型的类别 为空")
    return typeCategory



# 时间范围类型timeRangeType(支持输入: "created_date" "updated_date")，created_date 表示创建日期；而 updated_date 表示更新日期，支持输入为空
def get_timeRangeType():
    timeRangeType = input("请输入时间范围类型timeRangeType(支持输入: 'created_date' 'updated_date')，created_date 表示创建日期；而 updated_date 表示更新日期，支持输入为空: ")
    while timeRangeType not in ["created_date", "updated_date"] and len(timeRangeType) != 0:
        timeRangeType = input("请重输入时间范围类型timeRangeType(支持输入: 'created_date' 'updated_date')，created_date 表示创建日期；而 updated_date 表示更新日期，支持输入为空:")
    if len(timeRangeType) != 0:
        print(f"您输入的时间范围类型 是 {timeRangeType}")
    else:
        print(f"您输入的时间范围类型 为空")
    return timeRangeType


#输入3：统计维度groupBy(支持输入: "type" "type_category" "status" "status_category" )， type标准类型名称，type_category标准事务类别，status标准状态名称，status_category标准状态类别，当输入all时，代表查询全部，不支持输入为空
def get_issue_groupBy():
    issue_groupBy = input("请输入统计维度groupBy(支持输入: 'type' 'type_category' 'status' 'status_category')，type标准类型名称，type_category标准事务类别，status标准状态名称，status_category标准状态类别，当输入all时，代表查询全部，不支持输入为空: ")
    while issue_groupBy not in ["type", "type_category", "status", "status_category", "all"] or len(issue_groupBy) == 0:
        issue_groupBy = input("请重新输入统计维度groupBy(支持输入: 'type' 'type_category' 'status' 'status_category')，type标准类型名称，type_category标准事务类别，status标准状态名称，status_category标准状态类别，当输入all时，代表查询全部，不支持输入为空:")
    if issue_groupBy != 'all':
        print(f"您输入的 统计维度groupBy 是 {issue_groupBy}")
    else:
        print(f"您输入的是 'all', 稍后将查询全部统计维度groupBy: 'type' 'type_category' 'status' 'status_category'")
    return issue_groupBy


# 输入7：分组维度，支持输入 "sprint" "week" "month"，若不输入，则使用sprint
def get_groupDimension(default_groupDimension):
    groupDimension = input(f" 请输入分组维度，支持输入 'sprint' 'week' 'month'，若不输入，则使用{default_groupDimension}: ")
    while groupDimension not in ["sprint", "week", "month"] and len(groupDimension) != 0:
        groupDimension = input(f"请重新输入分组维度，支持输入 'sprint' 'week' 'month'，若不输入，则使用{default_groupDimension}:")
    if groupDimension:
        print(f"您输入的 统计维度groupBy 是 {groupDimension}")
    else:
        print(f"您输入的是 空, 稍后将使用 {default_groupDimension} 查询")
        groupDimension = default_groupDimension
    return groupDimension


# 2_Repo
# 保存输入匹配正则的代码库ID
def get_repo_id_pattern():
    special_repo_id_regex = input("请输入特定的代码库ID(支持正则匹配, 间隔符号仅支持英文), 不输入则查询所有代码库: ")
    repo_id_pattern = re.compile(special_repo_id_regex, re.IGNORECASE )
    # print(f"special_repo_id_regex = {special_repo_id_regex}\n")
    # print(f"repo_id_pattern = {repo_id_pattern}\n")
    if len(special_repo_id_regex) != 0:
        print(f"您输入的特定代码库ID 是 {special_repo_id_regex}")
    else:
        print(f"因为 您 输入的信息为空，所以此次接口调用，将查询项目中全部代码库")
    return special_repo_id_regex, repo_id_pattern



# 保存输入匹配正则的代码库ID或代码库名称
def get_repo_id_or_name_pattern():
    # 以下内容根据需要填写
    special_repo_regex = input("请输入特定的代码库ID 或 代码库名称(支持正则匹配, 间隔符号仅支持英文|) 或 all(代表查询全部代码库数据), 不支持输入为空: ")
    while len(special_repo_regex) == 0:
        special_repo_regex = input("请重新输入特定的代码库ID 或 代码库名称(支持正则匹配, 间隔符号仅支持英文|) 或 all(代表查询全部代码库数据), 不支持输入为空: ")

    repo_pattern = re.compile(special_repo_regex, re.IGNORECASE )

    if special_repo_regex != 'all':
        print(f"您输入的 特定的代码库(组)ID 或 代码库名称 是 {special_repo_regex}")
    else:
        print(f"您输入的是 'all', 稍后将查询全部代码库数据")

    return special_repo_regex, repo_pattern



# 保存输入匹配正则的代码库ID或代码库名称，不要求用户输入all
def get_repo_id_or_name_pattern_no_all():
    # 以下内容根据需要填写
    special_repo_regex = input("请输入特定的代码库ID 或 代码库名称(支持正则匹配, 间隔符号仅支持英文|) , 不支持输入为空: ")
    while len(special_repo_regex) == 0:
        special_repo_regex = input("请重新输入特定的代码库ID 或 代码库名称(支持正则匹配, 间隔符号仅支持英文|) , 不支持输入为空: ")

    repo_pattern = re.compile(special_repo_regex, re.IGNORECASE )

    print(f"您输入的 特定的代码库(组)ID 或 代码库名称 是 {special_repo_regex}")

    return special_repo_regex, repo_pattern




# 使用代码库ID 或 代码库名称 或 all 检索代码库
def find_repo_by_id_or_name(special_repo_regex, repo_pattern, group_list_dict):
    # 要查询的代码库的ids
    # repo_list = get_repo_list(client)
    # print(f"group_list_dict = {group_list_dict}")
    # list to dict
    repo_dict = {}

    if special_repo_regex == 'all':
        for index, i in enumerate(group_list_dict.items()):
            repo_dict[i[1]['id']] = {}
            repo_dict[i[1]['id']] = i[1]
            repo_dict[i[1]['id']]['代码库ID'] = i[1]['id']
            repo_dict[i[1]['id']]['检索命中字段'] = "列出全部代码库"
            repo_dict[i[1]['id']]['被检索字段'] = "列出全部代码库"
            repo_dict[i[1]['id']]['代码库名称'] = i[1]['name']

    else:
        for index_2, i_2 in enumerate(group_list_dict.items()):
            if repo_pattern.findall(i_2[1]['id']):
                repo_dict[i_2[1]['id']] = {}
                repo_dict[i_2[1]['id']] = i_2[1]
                repo_dict[i_2[1]['id']]['代码库ID'] = i_2[1]['id']
                repo_dict[i_2[1]['id']]['检索命中字段'] = '代码库ID'

            if repo_pattern.findall(i_2[1]['name']):
                if i_2[1]['id'] in repo_dict.keys():
                    origin_str = repo_dict[i_2[1]['id']]['检索命中字段']
                    repo_dict[i_2[1]['id']]['检索命中字段'] = f"{origin_str} & 代码库名称"
                else:
                    repo_dict[i_2[1]['id']] = {}
                    repo_dict[i_2[1]['id']] = i_2[1]
                    repo_dict[i_2[1]['id']]['代码库ID'] = i_2[1]['id']
                    repo_dict[i_2[1]['id']]['检索命中字段'] = '代码库名称'

            if repo_pattern.findall(i_2[1]['id']) or repo_pattern.findall(i_2[1]['name']):
                repo_dict[i_2[1]['id']]['被检索字段'] = special_repo_regex
                repo_dict[i_2[1]['id']]['代码库名称'] = i_2[1]['name']



    if len(repo_dict) == 0:
        print(f"!!!!!!很抱歉, 系统没有找到匹配代码库ID、代码库名称 正则表达式的: {special_repo_regex} 的代码库, 可能此项目下无代码库，或者请您重新执行脚本文件, 并修改代码库信息, 以便找到您需要的代码库")

    print(f"本次检索到的代码库数量是: {len(repo_dict.keys())}")
    # print(f"/n")
    # print(f"repo_dict = {repo_dict}")
    return repo_dict



# 保存输入匹配正则的分支名称
def get_branch_pattern():
    # 以下内容根据需要填写
    special_branch_regex = input("请输入特定的分支名称(支持正则匹配, 间隔符号仅支持英文|) 或 HEAD(代表查询代码库默认分支) 或 all(代表查询全部分支数据), 不支持输入为空: ")
    while len(special_branch_regex) == 0:
        special_branch_regex = input("请重新输入特定的 分支名称(支持正则匹配, 间隔符号仅支持英文|) 或 HEAD(代表查询代码库默认分支) 或 all(代表查询全部分支数据), 不支持输入为空: ")

    branch_pattern = re.compile(special_branch_regex, re.IGNORECASE )

    if special_branch_regex != 'all':
        print(f"您输入的 特定的分支名称 是 {special_branch_regex}")
    else:
        print(f"您输入的是 'all', 稍后将查询全部分支数据")

    return special_branch_regex, branch_pattern








# 输入pageSize, 如果输入为空, 则在调接口时, 不使用此参数, 输入不为空时仅支持输入非0数字
def get_pageSize(default_pageSize):
    pageSize = input(f"请输入您想检索的每页记录数, 不输入则使用默认值: ")
    while is_valid_number(pageSize) and len(pageSize) != 0:
        pageSize = input(f"请重新输入您想检索的每页记录数, 不输入则使用默认值: ")
    if len(pageSize) != 0:
        print(f"您输入的 每页记录数 是 {pageSize}")
        try:
            pageSize = int(pageSize)
        except Exception as e:
            print(f"Error : {e}")
            pageSize = default_pageSize
    else:
        print(f"您输入的 每页记录数 为空, 稍后将按照默认的每页记录数{default_pageSize} 检索")
        pageSize = default_pageSize
    return pageSize



# 保存输入匹配正则的项目ID
def get_project_id_pattern():
    special_project_id_regex = input("请输入特定的项目(组)ID(支持正则匹配, 间隔符号仅支持英文|), 不输入则查询所有项目: ")
    project_id_pattern = re.compile(special_project_id_regex, re.IGNORECASE )
    # print(f"special_project_id_regex = {special_project_id_regex}\n")
    # print(f"project_id_pattern = {project_id_pattern}\n")

    if len(special_project_id_regex) != 0:
        print(f"您输入的 特定的项目(组)ID 是 {special_project_id_regex}")
    else:
        print(f"您输入的 特定的项目(组)ID 为空, 稍后将查询全部项目(组)数据")

    return special_project_id_regex, project_id_pattern



# 保存输入匹配正则的项目ID或项目名称
def get_project_id_or_name_pattern():
    # 以下内容根据需要填写
    special_project_regex = input("请输入特定的项目ID 或 项目名称(支持正则匹配, 间隔符号仅支持英文|) 或 all(代表查询全部项目), 不支持输入为空: ")
    while len(special_project_regex) == 0:
        special_project_regex = input("请重新输入特定的项目ID 或 项目名称(支持正则匹配, 间隔符号仅支持英文|) 或 all(代表查询全部项目), 不支持输入为空: ")

    project_pattern = re.compile(special_project_regex, re.IGNORECASE )

    if len(special_project_regex) != 'all':
        print(f"您输入的 特定的项目(组)ID 或 项目名称 是 {special_project_regex}")
    else:
        print(f"您输入的是 'all', 稍后将查询全部项目(组)数据")

    return special_project_regex, project_pattern



# 使用项目ID 或 项目名称 或 all 检索项目
def find_project_by_id_or_name(special_project_regex, project_pattern, group_list_dict):
    # 要查询的项目的ids
    # project_list = get_project_list(client)
    # print(f"group_list_dict = {group_list_dict}")
    # list to dict
    project_dict = {}

    if special_project_regex == 'all':
        for index, i in enumerate(group_list_dict.items()):
            project_dict[i[1]['id']] = {}
            project_dict[i[1]['id']] = i[1]
            project_dict[i[1]['id']]['项目ID'] = i[1]['id']
            project_dict[i[1]['id']]['检索命中字段'] = "列出全部项目"
            project_dict[i[1]['id']]['被检索字段'] = "列出全部项目"
            project_dict[i[1]['id']]['项目名称'] = i[1]['name']

    else:
        for index_2, i_2 in enumerate(group_list_dict.items()):
            if project_pattern.findall(i_2[1]['id']):
                project_dict[i_2[1]['id']] = {}
                project_dict[i_2[1]['id']] = i_2[1]
                project_dict[i_2[1]['id']]['项目ID'] = i_2[1]['id']
                project_dict[i_2[1]['id']]['检索命中字段'] = '项目ID'

            if project_pattern.findall(i_2[1]['name']):
                if i_2[1]['id'] in project_dict.keys():
                    origin_str = project_dict[i_2[1]['id']]['检索命中字段']
                    project_dict[i_2[1]['id']]['检索命中字段'] = f"{origin_str} & 项目名称"
                else:
                    project_dict[i_2[1]['id']] = {}
                    project_dict[i_2[1]['id']] = i_2[1]
                    project_dict[i_2[1]['id']]['项目ID'] = i_2[1]['id']
                    project_dict[i_2[1]['id']]['检索命中字段'] = '项目名称'

            if project_pattern.findall(i_2[1]['id']) or project_pattern.findall(i_2[1]['name']):
                project_dict[i_2[1]['id']]['被检索字段'] = special_project_regex
                project_dict[i_2[1]['id']]['项目名称'] = i_2[1]['name']



    if len(project_dict) == 0:
        print(f"!!!!!!很抱歉, 系统没有找到匹配项目ID、项目名称 正则表达式的: {special_project_regex} 的项目, 请您重新执行脚本文件, 并修改项目信息, 以便找到您需要的项目")

    print(f"目前累计检索到的项目数量是: {len(project_dict.keys())}")
    # print(f"/n")
    # print(f"project_dict = {project_dict}")
    return project_dict




# 保存使用哪种项目统计模式：
# 仅子节点(child nodes only)或者全部后代节点(all descendant nodes)，输入1或者2进行选择
# 若不输入，则默认使用 1 仅子节点模式(child nodes only)
def get_project_statistical_mode():
    special_project_statistical_mode = input("请确定使用哪一种项目统计模式：仅子节点(child nodes only)或者全部后代节点(all descendant nodes)，输入1或者2进行选择，若不输入，则默认使用 1 仅子节点(child nodes only)模式: ")
    while special_project_statistical_mode not in ["1", "2"] and len(special_project_statistical_mode) != 0:
        special_project_statistical_mode = input("请重新确定使用哪一种项目统计模式：仅子节点(child nodes only)或者全部后代节点(all descendant nodes)，输入1或者2进行选择，若不输入，则默认使用 1 仅子节点(child nodes only)模式: ")

    if special_project_statistical_mode == "1" or special_project_statistical_mode == "":
        special_project_statistical_mode_number = "1"
        special_project_statistical_mode = "child nodes only"
    elif special_project_statistical_mode == "2":
        special_project_statistical_mode_number = "2"
        special_project_statistical_mode = "all descendant nodes"

    # print(f"special_project_statistical_mode_number = {special_project_statistical_mode_number}\n")
    # print(f"special_project_statistical_mode = {special_project_statistical_mode}\n")
    return special_project_statistical_mode_number, special_project_statistical_mode



# 保存使用哪种repo 依赖包 分组依据，package或者tag(标签)
# 输入1或者2进行选择，若不输入，则默认使用 1 package分组
def get_group_by():
    special_group_by = input("请确定使用哪一种分组依据：package或者tag，输入1或者2进行选择，若不输入，则默认使用 1 package依据: ")
    while special_group_by not in ["1", "2"] and len(special_group_by) != 0:
        special_group_by = input("请重新确定使用哪一种分组依据：package或者tag，输入1或者2进行选择，若不输入，则默认使用 1 package依据: ")

    if special_group_by == "1" or special_group_by == "":
        special_group_by_number = "1"
        special_group_by = "package"
    elif special_group_by == "2":
        special_group_by_number = "2"
        special_group_by = "tag"

    # print(f"special_group_by_number = {special_group_by_number}\n")
    # print(f"special_group_by = {special_group_by}\n")
    return special_group_by_number, special_group_by



# 保存输入匹配正则的开发者邮箱
def get_developer_email_pattern():
    special_developer_email_regex = input("请输入特定的开发者邮箱(支持正则匹配, 间隔符号仅支持英文|), 若不输入，则查询全部开发者的数据: ")
    developer_email_pattern = re.compile(special_developer_email_regex, re.IGNORECASE )
    # print(f"special_developer_email_regex = {special_developer_email_regex}\n")
    # print(f"developer_email_pattern = {developer_email_pattern}\n")
    return special_developer_email_regex, developer_email_pattern



# 保存输入的贡献者提交name列表，填写多个name时，间隔符号用支持英文逗号，可以不输入，当输入为空时，默认就传空，代表查询全部贡献者
def get_developer_name_list():
    developer_name_list = input("请输入贡献者提交name列表，填写多个name时，间隔符号用支持英文逗号，可以不输入，当输入为空时，默认就传空，代表查询全部贡献者: ")
    return developer_name_list



# 保存输入的指定一个开发者邮箱，如果不输入则表示直接查询全部开发者的数据
def get_single_developer_email():
    single_developer_email = input("请输入指定一个开发者邮箱，如果不输入则表示直接查询全部开发者的数据: ")
    return single_developer_email



# 保存输入的指定开发者邮箱列表，如果不输入则表示直接查询全部开发者的数据
def get_developer_email_list():
    developer_email_list = input("请输入指定的开发者邮箱列表，如果不输入则表示直接查询全部开发者的数据: ")
    return developer_email_list



# 把refs/remotes/origin/reports,refs/tags/v2.0.0,HEAD 把这几个远端分支的分支名称过滤一下，只保留 origin/后面的内容， 或者只保留 tags/后面的内容，或者直接返回HEAD
def filter_refs(refs):
    # 过滤出 origin/ 和 tags/ 后面的内容
    filtered_refs = []
    for ref in refs:
        if ref == 'HEAD':
            filtered_refs.append(ref)  # 如果是 'HEAD'，直接添加到列表
        else:
            # 匹配 origin/ 和 tags/ 后面的内容
            origin_match = re.match(r"refs/remotes/origin/(.*)", ref)
            tags_match = re.match(r"refs/tags/(.*)", ref)
            if origin_match:
                filtered_refs.append(origin_match.group(1))
            elif tags_match:
                filtered_refs.append(tags_match.group(1))
    return filtered_refs



# 保存开始日期，支持传空
def get_start_date_allow_empty(days_count):
    while is_valid_number(days_count):
        days_count = input("请重新输入开始日期的向前数的日期: ")
    date_days_count_ago = datetime.strftime(datetime.now() - timedelta(days=days_count), "%Y-%m-%d")
    start_date = input(f"请输入查询开始日期，例如{date_days_count_ago}: ")
    while is_invalid_date(start_date) and len(start_date) != 0:
        start_date = input(f"请重新输入符合年月日规则的日期， 例如{date_days_count_ago}: ")
    if len(start_date) == 0:
        print(f"您输入的开始日期 是 空 ")
    else:
        print(f"您输入的开始日期 是 {start_date}")
    return start_date



# 保存开始日期，支持传空
def get_start_date_allow_empty_by_date_format(days_count):
    date_days_count_ago = datetime.strftime(datetime.now() - timedelta(days=days_count), "%Y-%m-%d")
    start_date = input(f"请输入查询开始日期，例如{date_days_count_ago}: ")
    while is_invalid_date(start_date) and len(start_date) != 0:
        start_date = input(f"请重新输入符合年月日规则的日期，例如{date_days_count_ago}: ")
    if len(start_date) == 0:
        print(f"您输入的开始日期 是 空 ")
    else:
        print(f"您输入的开始日期 是 {start_date}")
    return start_date




# 保存结束日期，支持传空
def get_end_date_allow_empty():
    date_now = datetime.strftime(datetime.now(), "%Y-%m-%d")
    end_date = input(f"请输入查询结束日期，例如{date_now}: ")
    while is_invalid_date(end_date) and len(end_date) != 0:
        end_date = input(f"请重新输入符合年月日规则的日期，例如{date_now}: ")
    if len(end_date) == 0:
        print(f"您输入的结束日期 是 空 ")
    else:
        print(f"您输入的结束日期 是 {end_date}")
    return end_date


# 通用方法
# 保存用户输入的一个支持传空的普通日期，支持输入为空
def get_normal_date_allow_empty(date_name):
    date_now = datetime.strftime(datetime.now(), "%Y-%m-%d")
    normal_date = input(f"请输入 {date_name}，例如{date_now}: ")
    while is_invalid_date(normal_date) and len(normal_date) != 0:
        normal_date = input(f"请重新输入符合年月日规则的日期，例如{date_now}: ")
    if len(normal_date) == 0:
        print(f"您输入的{date_name} 是 空 ")
    else:
        print(f"您输入的{date_name} 是 {normal_date}")
    return normal_date



# 输入时间量值, 单位为 week
def get_unitNumber(example):
    unitNumber = input(f"请输入时间量值，例如{example}: ")
    while is_valid_number(unitNumber) and len(unitNumber) != 0:
        unitNumber = input(f"请重新输入符合规则的时间量值，例如{example}: ")
    if len(unitNumber) == 0:
        print(f"您输入的时间量值 是 空 , 稍后将使用 1 作为默认时间量值")
    else:
        print(f"您输入的时间量值 是 {unitNumber}")
    return unitNumber



# 输入行业指标类型, Enum: "static_test_coverage" "doc_coverage" "modularity" "dryness" "issue_rate" "severe_issue_rate"
def get_metricType():
    metricType = input(f"请输入行业指标类型，可以输入包括static_test_coverage、doc_coverage、modularity、dryness、\
issue_rate、severe_issue_rate，若不输入直接查询全部行业指标类型: ")
    while not(metricType in ['static_test_coverage', 'doc_coverage', 'modularity', 'dryness',\
     'issue_rate', 'severe_issue_rate']) and len(metricType) != 0:
        metricType = input(f"请重新输入行业指标类型，可以输入包括static_test_coverage、doc_coverage、modularity、dryness、\
issue_rate、severe_issue_rate，若不输入直接查询全部行业指标类型: ")

    if len(metricType) == 0:
        print(f"您输入的行业指标类型 为 空 ，稍后将查询全部行业指标类型的数据")
    else:
        print(f"您输入的行业指标类型 为 {metricType}")

    return metricType




# 保存输入的开始日期
def get_start_date(days_count):
    while is_valid_number(days_count):
        days_count = input("请重新输入开始日期的向前数的日期: ")
    date_year_ago = datetime.strftime(datetime.now() - timedelta(days=days_count), "%Y-%m-%d")

    start_date = input(f"请输入查询开始日期，例如{date_year_ago}，若不输入则使用{date_year_ago}: ")
    while is_invalid_date(start_date) and len(start_date) != 0:
        start_date = input(f"请重新输入符合年月日规则的日期， 例如{date_year_ago}，若不输入则使用{date_year_ago}: ")

    if len(start_date) == 0:
        start_date = date_year_ago

    print(f"开始日期 是 {start_date}")
    return start_date



# 保存输入的结束日期
def get_end_date():
    date_now = datetime.strftime(datetime.now(), "%Y-%m-%d")

    end_date = input(f"请输入查询结束日期，例如{date_now}，若不输入则使用{date_now}: ")
    while is_invalid_date(end_date) and len(end_date) != 0:
        end_date = input(f"请重新输入符合年月日规则的日期， 例如{date_now}，若不输入则使用{date_now}: ")

    if len(end_date) == 0:
        end_date = date_now

    print(f"结束日期 是 {end_date}")
    return end_date


# 把开始和结束日期从date改为date-time形式
def generate_total_datetime(start_date, end_date):
    start = datetime.strptime(start_date, "%Y-%m-%d")
    start = start - timedelta(hours=8)
    start = datetime.strftime(start,"%Y-%m-%dT%H:%M:%S.000Z")

    end = datetime.strptime(end_date, "%Y-%m-%d")
    end = end + timedelta(hours=16)
    end = datetime.strftime(end,"%Y-%m-%dT%H:%M:%S.000Z")

    return start, end


# 保存输入是否需要每个代码库的总代码当量数据：Enum: "true" "false"，若不输入则为false
def get_whether_need_sum_efficiency_metrics():
    result = input(f"是否需要每个代码库的总代码当量数据，可以输入true或者输入false，若不输入则为false: ")
    while len(result) != 0 and not(result in ['true', 'false']):
        result = input(f"请重新输入是否需要每个代码库的总代码当量数据，可以输入true或者输入false，若不输入则为false:")
    if result == 'true':
        result = True
    elif result == 'false' or len(result) == 0:
        result = False
    print(f"是否需要每个代码库的总代码当量数据 是 {result}")
    return result


# 保存输入的需要聚合的时间范围
def get_unit_of_time(unit_of_time):
    unitOfTime = input(f"请输入统计步长（需要聚合的时间范围），可以输入包括day、week、month、quarter、year，若不输入则使用{unit_of_time}: ")
    while not(unitOfTime in ['day', 'week', 'month', 'quarter', 'year']) and len(unitOfTime) != 0:
        unitOfTime = input(f"请重新输入统计步长（需要聚合的时间范围），可以输入包括day、week、month、quarter、year，若不输入则使用{unit_of_time}: ")
    if len(unitOfTime) == 0:
        unitOfTime = unit_of_time
    print(f"统计步长（需要聚合的时间范围） 是 {unitOfTime}")
    return unitOfTime



# 保存输入的需要聚合的时间范围
def get_unit_of_time_general(unit_of_time, date_list):
    unitOfTime = input(f"请输入统计步长（需要聚合的时间范围），可以输入包括{date_list}，若不输入则使用{unit_of_time}: ")
    while not(unitOfTime in date_list) and len(unitOfTime) != 0:
        unitOfTime = input(f"请重新输入统计步长（需要聚合的时间范围），可以输入包括{date_list}，若不输入则使用{unit_of_time}: ")
    if len(unitOfTime) == 0:
        unitOfTime = unit_of_time
    print(f"统计步长（需要聚合的时间范围） 是 {unitOfTime}")
    return unitOfTime



# 保存输入的需要聚合的时间范围
def get_unit_of_time_allow_empty():
    unitOfTime = input(f"请输入统计步长（需要聚合的时间范围），可以输入包括day、week、month、quarter、year，若不输入直接查询总和数据，不按时间步长分割数据: ")
    while not(unitOfTime in ['day', 'week', 'month', 'quarter', 'year']) and len(unitOfTime) != 0:
        unitOfTime = input(f"请重新输入统计步长（需要聚合的时间范围），可以输入包括day、week、month、quarter、year，若不输入直接查询总和数据，不按时间步长分割数据: ")

    if len(unitOfTime) == 0:
        print(f"您输入的统计步长（需要聚合的时间范围）为 空 ，稍后将只查询总和数据")
    else:
        print(f"您输入的统计步长（需要聚合的时间范围）为 {unitOfTime}")

    return unitOfTime



# 保存输入的统计时间步长的数值
def get_unit_of_number(unit_of_number):
    unitOfNumber = input(f"请输入统计时间步长的数值，可以输入任意正整数，若不输入则使用{unit_of_number}: ")
    while not is_positive_integer(unitOfNumber) and len(unitOfNumber) != 0:
        unitOfNumber = input(f"请重新输入统计时间步长的数值，可以输入任意正整数，若不输入则使用{unit_of_number}: ")

    if len(unitOfNumber) == 0:
        unitOfNumber = unit_of_number

    print(f"统计时间步长的数值 是 {unitOfNumber}")
    return unitOfNumber



# 判断一个输入是否是正整数，如果是正整数就返回True，否则返回False
def is_positive_integer(s):
    try:
        num = int(s)
        if num > 0:
            if str(num) == s:  # 检查是否包含其他非数字字符
                return True
            else:
                return False
        else:
            return False
    except ValueError:
        return False


# 判断一个输入是否是自然数，如果是自然数就返回True，否则返回False
def is_natural_number(s):
    try:
        num = int(s)
        if num >= 0:
            if str(num) == s:  # 检查是否包含其他非数字字符
                return True
            else:
                return False
        else:
            return False
    except ValueError:
        return False



# 保存客户输入的时区
def get_time_zone_Name(time_zone_Name):
    timezoneName = input(f"请输入时区缩写，若不输入则使用{time_zone_Name}(东八区): ")

    while not is_valid_timezone(timezoneName) and len(timezoneName) != 0:
        timezoneName = input(f"请重新输入一个有效的时区缩写，若不输入则使用{time_zone_Name}(东八区): ")

    if len(timezoneName) == 0:
        timezoneName = time_zone_Name

    print(f"您输入的 时区缩写 是 {timezoneName}")
    return timezoneName



# 判断是否是一个有效时区，如果是有效时区返回True，否则返回False
def is_valid_timezone(timezone):
    try:
        pytz.timezone(timezone)
        return True
    except pytz.exceptions.UnknownTimeZoneError:
        return False



# 保存执行模式，1为集中执行，2为分散执行，默认为1
def get_execution_mode():
    result = input(f"请输入您执行查询时的执行模式，1为完全集中(在一次接口调用中全部传参)，2为分批集中(在一次接口中传一批参数)，3为分散传参(在一次接口调用中仅传递一个参数)，不输入则为集中执行: ")
    while len(result) != 0 and result not in ["1","2","3"]:
        result = input(f"请重新输入您执行查询时的执行模式，1为完全集中(在一次接口调用中全部传参)，2为分批集中(在一次接口中传一批参数)，3为分散传参(在一次接口调用中仅传递一个参数)，不输入则为集中执行: ")

    if len(result) == 0:
        result = "1"
    print(f"您输入的执行模式为 {result}, 1为完全集中(在一次接口调用中全部传参)，2为分批集中(在一次接口中传一批参数)，3为分散传参(在一次接口调用中仅传递一个参数)")
    return result
    


# 将批量传参的参数，分批执行
# def chunk_dict(data_dict, chunk_size):
#     # 获取字典的键列表
#     keys = list(data_dict.keys())
#     # 创建一个空列表用于存储分割后的子字典
#     chunked_dicts = []
#     # 使用range()函数来按chunk_size步长遍历键列表
#     for i in range(0, len(keys), chunk_size):
#         # 获取当前chunk的键
#         chunk_keys = keys[i:i + chunk_size]
#         # 创建一个子字典，并将当前chunk的键值对添加进去
#         chunk_dict = {key: data_dict[key] for key in chunk_keys}
#         # 将子字典添加到chunked_dicts列表中
#         chunked_dicts.append(chunk_dict)
#     return chunked_dicts
def chunk_dict(data, chunk_size):
    # 如果输入是字典
    if isinstance(data, dict):
        # 获取字典的键列表
        keys = list(data.keys())
        # 创建一个空列表用于存储分割后的子字典
        chunked_dicts = []
        # 使用range()函数来按chunk_size步长遍历键列表
        for i in range(0, len(keys), chunk_size):
            # 获取当前chunk的键
            chunk_keys = keys[i:i + chunk_size]
            # 创建一个子字典，并将当前chunk的键值对添加进去
            chunk_dict = {key: data[key] for key in chunk_keys}
            # 将子字典添加到chunked_dicts列表中
            chunked_dicts.append(chunk_dict)
        return chunked_dicts

    # 如果输入是列表
    elif isinstance(data, list):
        # 使用列表切片将列表分割成指定大小的块
        chunked_lists = [data[i:i + chunk_size] for i in range(0, len(data), chunk_size)]
        return chunked_lists

    # 如果输入既不是字典也不是列表，抛出类型错误
    else:
        raise TypeError("Input data must be a dictionary or a list.")




# 保存是否仅查询测试/非测试代码的当量 (此过滤条件仅支持开发当量，开发价值指标)
def get_test_code():
    testCode_result = None
    testCode = input(f"是否仅查询测试/非测试代码的当量 (此过滤条件仅支持开发当量，开发价值指标)，可以不输入或者输入true或者输入false: ")
    while len(testCode) != 0 and not(testCode in ['true', 'false']):
        testCode = input(f"请重新输入是否仅查询测试/非测试代码的当量 (此过滤条件仅支持开发当量，开发价值指标)，可以不输入或者输入true或者输入false: ")

    if testCode == 'true':
        testCode_result = True
    elif testCode == 'false':
        testCode_result = False

    print(f"是否仅查询测试/非测试代码的当量 是 {testCode_result}")
    return testCode_result


# 保存输入的包含的目录列表
def get_includePath():
    includePath_regex = input("请输入包含的目录列表（相对项目根目录的文件路径前缀）includePath，若不输入，则表示包含所有目录: ")

    if len(includePath_regex) != 0:
        print(f"包含的目录列表（相对项目根目录的文件路径前缀）是 {includePath_regex}")
    else:
        print(f"因为 您 输入的包含目录列表为空，所以此次接口调用 包含所有目录")
    return includePath_regex


# 保存输入的排除的目录列表
def get_excludePath():
    excludePath_regex = input("请输入排除的目录列表（相对项目根目录的文件路径前缀）excludePath，若不输入，则表示不排除任何目录: ")

    if len(excludePath_regex) != 0:
        print(f"排除的目录列表（相对项目根目录的文件路径前缀）是 {excludePath_regex}")
    else:
        print(f"因为 您 输入的排除目录列表为空，所以此次接口调用 不排除任何目录")
    return excludePath_regex


# 保存输入匹配正则的start_hash
def get_start_hash_pattern():
    special_start_hash_regex = input("请输入特定的开始hash（包括）(支持正则匹配)，若不输入，则默认使用通过思码逸代码库分析统计到的该代码库最早提交的commit hash: ")

    start_hash_pattern = re.compile(special_start_hash_regex, re.IGNORECASE )
    # print(f"special_start_hash_regex = {special_start_hash_regex}\n")
    # print(f"start_hash_pattern = {start_hash_pattern}\n")
    if len(special_start_hash_regex) != 0:
        print(f"特定的开始hash（包括）(支持正则匹配) 是 {special_start_hash_regex}")
    else:
        print(f"因为 您 输入的特定的开始hash为空，所以此次接口调用默认使用通过思码逸代码库分析统计到的该代码库最早提交的commit hash ")
    return special_start_hash_regex, start_hash_pattern


# 保存输入匹配正则的end_hash
def get_end_hash_pattern():
    special_end_hash_regex = input("请输入特定的结束hash（包括）(支持正则匹配)，若不输入，则默认使用通过思码逸代码库分析统计到的该代码库最晚提交的commit hash: ")

    end_hash_pattern = re.compile(special_end_hash_regex, re.IGNORECASE )
    # print(f"special_end_hash_regex = {special_end_hash_regex}\n")
    # print(f"end_hash_pattern = {end_hash_pattern}\n")
    print(f"特定的结束hash（包括）(支持正则匹配) 是 {special_end_hash_regex}")
    return special_end_hash_regex, end_hash_pattern


# 保存输入 jira issue id
def get_jira_issue_id_pattern():
    special_jira_issue_id_regex = input("请输入特定的 jira issue ID(支持正则匹配, 间隔符号仅支持英文), 不输入则查询全部: ")
    jira_issue_id_pattern = re.compile(special_jira_issue_id_regex, re.IGNORECASE )
    # print(f"special_jira_issue_id_regex = {special_jira_issue_id_regex}\n")
    # print(f"jira_issue_id_pattern = {jira_issue_id_pattern}\n")
    return special_jira_issue_id_regex, jira_issue_id_pattern



# 保存输入的时间周期数量 recentPeriod_counts
def get_recentPeriod_counts():
    recentPeriod_counts = input("请输入您想检索的回溯时间参数---时间周期数量: ")
    while is_valid_number(recentPeriod_counts) and len(recentPeriod_counts) != 0:
        recentPeriod_counts = input("请重新输入您想检索的回溯时间参数---时间周期数量: ")
    if len(recentPeriod_counts) != 0:
        print(f"您输入的 回溯时间参数---检索时间周期数量 是 {recentPeriod_counts}")
    else:
        print(f"您输入的 回溯时间参数---检索时间周期数量 为空")
    return recentPeriod_counts



# 输入时间周期单位 recentPeriod_unit
def get_recentPeriod_unit():
    recentPeriod_unit = input("请输入您想检索的回溯时间参数---时间周期单位，支持'day','week','month','quarter','year' 或不输入: ")
    while  recentPeriod_unit not in ["day","week","month","quarter","year"] and len(recentPeriod_unit) != 0:
        recentPeriod_unit = input(f"请重新输入您想检索的回溯时间参数---时间周期单位，支持'day','week','month','quarter','year' 或不输入: ")
    if len(recentPeriod_unit) != 0:
        print(f"您输入的 回溯时间参数---检索时间周期单位 是 {recentPeriod_unit}")
    else:
        print(f"您输入的 回溯时间参数---检索时间周期单位 为空")
    return recentPeriod_unit


# get_since
def get_since():
    date_day_ago = datetime.strftime(datetime.now() - timedelta(days=1), "%Y-%m-%d")
    since = input(f"请输入有效的since 起始日期，例如{date_day_ago}: ")

    while is_invalid_date(since) and len(since) != 0:
        print(f"since = {since}")
        since = input(f"请重新输入符合年月日规则的since 起始日期， 例如{date_day_ago}: ")

    if len(since) != 0:
        print(f"您输入的 since 起始日期 是 {since}")
    else:
        print(f"您输入的 since 起始日期 为空")
    return since


# get_until
def get_until():
    date_day_ago = datetime.strftime(datetime.now() - timedelta(days=1), "%Y-%m-%d")
    until = input(f"请输入有效的until 结束日期，例如{date_day_ago}: ")

    while is_invalid_date(until) and len(until) != 0:
        print(f"until = {until}")
        until = input(f"请重新输入符合年月日规则的until 结束日期， 例如{date_day_ago}: ")

    if len(until) != 0:
        print(f"您输入的 until 结束日期 是 {until}")
    else:
        print(f"您输入的 until 结束日期 为空")
    return until



# 2_repo_query-quality-metrics-graph
# 输入时间范围 的开始日期
def get_timeRange_of_start():
    date_day_ago = datetime.strftime(datetime.now() - timedelta(days=1), "%Y-%m-%d")
    timeRange_of_start = input(f"请输入时间范围 的开始日期，例如{date_day_ago}: ")

    while is_invalid_date(timeRange_of_start) and len(timeRange_of_start) != 0:
        print(f"timeRange_of_start = {timeRange_of_start}")
        timeRange_of_start = input(f"请重新输入符合年月日规则的开始日期， 例如{date_day_ago}: ")

    if len(timeRange_of_start) != 0:
        print(f"时间范围 的开始日期 是 {timeRange_of_start}")
    else:
        print(f"时间范围 的开始日期 是 空")
    return timeRange_of_start


# 2_repo_query-quality-metrics-graph
# 输入时间范围 的结束日期
def get_timeRange_of_end():
    date_now = datetime.strftime(datetime.now(), "%Y-%m-%d")
    timeRange_of_end = input(f"请输入时间范围 的结束日期，例如{date_now}: ")
    while is_invalid_date(timeRange_of_end) and len(timeRange_of_end) != 0:
        timeRange_of_end = input(f"请重新输入符合年月日规则的结束日期， 例如{date_now}: ")

    if len(timeRange_of_end) != 0:
        print(f"时间范围 的结束日期 是 {timeRange_of_end}")
    else:
        print(f"时间范围 的开始日期 是 空")
    return timeRange_of_end


# 2_repo_query-quality-metrics-graph
# 输入查询类型，项目或者代码库，project or repository
def get_records_type():
    records_type = input("请输入您想检索的查询类型，支持'project','repository': ")
    while records_type not in ["project","repository"]:
        records_type = input(f"请重新输入您想检索的查询类型，支持'project','repository': ")

    if len(records_type) != 0:
        print(f"您输入的 检索的查询类型 是 {records_type}")
    else:
        print(f"您输入的 检索的查询类型 为空")
    return records_type


# 2_repo_query-quality-metrics-graph
# 输入项目ID或者代码库ID，不输入则查询所有项目或所有代码库
def get_records_id():
    records_id = input("请项目ID或者代码库ID，不输入则查询所有项目或所有代码库: ")

    if len(records_id) != 0:
        print(f"您输入的 项目ID或者代码库ID 是 {records_id}")
    else:
        print(f"您输入的 项目ID或者代码库ID 为空，稍后将会查询所有项目或所有代码库")
    return records_id

# 2_repo_query-quality-metrics-graph
# 如果查询所有项目或所有代码库，则输入的代码库ID列表数组自动失效
# 只有查单个项目或单个代码库时，输入代码库ID列表才有效
# 输入代码库ID列表
def get_repoIds():
    repoIds = input("请输入代码库ID列表: ")

    if len(repoIds) != 0:
        repoIds = repoIds.split('|')
        print(f"您输入的 代码库ID列表 是 {repoIds}")
    else:
        print(f"您输入代码库ID列表 为空")
    return repoIds

# 2_repo_query-quality-metrics-graph
# 如果查询所有项目或所有代码库，则输入的贡献者列表数组自动失效
# 只有查单个项目或单个代码库时，输入贡献者列表数组才有效
# 输入贡献者列表
def get_contributors():
    contributors = input("请输入贡献者列表: ")

    if len(contributors) != 0:
        contributors = contributors.split('|')
        print(f"您输入的 贡献者列表 是 {contributors}")
    else:
        print(f"您输入的 贡献者列表 为空")
    return contributors



# 2_repo_12_repo_commit_list_by_commit_label
# 根据标签组查询commit指标
# 输入标签组名：groupNames
def get_groupNames():
    groupNames = input("请输入标签组名: ")

    if len(groupNames) != 0:
        groupNames = groupNames.split('|')
        print(f"您输入的 标签组名 是 {groupNames}")
    else:
        print(f"您输入的 标签组名 为空")
    return groupNames



# 2_repo_12_repo_commit_list_by_commit_label
# 根据标签组查询commit指标
# 输入标签：labelValues
def get_labelValues():
    labelValues = input("请输入标签: ")

    if len(labelValues) != 0:
        labelValues = labelValues.split('|')
        print(f"您输入的 标签 是 {labelValues}")
    else:
        print(f"您输入的 标签 为空")
    return labelValues



# 2_repo_12_repo_commit_list_by_commit_label
# 根据标签组查询commit指标
# 输入Commit起始时间：commitStartTime
def get_commitStartTime():
    datetime_day_ago = datetime.strftime(datetime.now() - timedelta(days=7), "%Y-%m-%d %H:%M:%S")
    commitStartTime = input(f"请输入有效的Commit创建的起始时间，例如{datetime_day_ago}: ")

    while is_invalid_datetime(commitStartTime) and len(commitStartTime) != 0:
        commitStartTime = input(f"请重新输入符合年月日时分秒规则的日期时间，例如{datetime_day_ago}: ")

    if len(commitStartTime) != 0:
        print(f"您输入的 Commit 创建的起始时间 是 {commitStartTime}")
        commitStartTime = prc_to_utc_time(commitStartTime)
    else:
        print(f"您输入的 Commit 创建的起始时间 为空")
    return commitStartTime



# 2_repo_12_repo_commit_list_by_commit_label
# 根据标签组查询commit指标
# 输入Commit截止时间：commitEndTime
def get_commitEndTime():
    datetime_day_ago = datetime.strftime(datetime.now() - timedelta(days=1), "%Y-%m-%d %H:%M:%S")
    commitEndTime = input(f"请输入有效的Commit创建的截止时间，例如{datetime_day_ago}: ")

    while is_invalid_datetime(commitEndTime) and len(commitEndTime) != 0:
        commitEndTime = input(f"请重新输入符合年月日时分秒规则的日期时间，例如{datetime_day_ago}: ")

    if len(commitEndTime) != 0:
        print(f"您输入的 Commit 创建的截止时间 是 {commitEndTime}")
        commitEndTime = prc_to_utc_time(commitEndTime)
    else:
        print(f"您输入的 Commit 创建的截止时间 为空")
    return commitEndTime



# 14_CodeIssue_01_code_issue_list
# 获取提交者Emails，一次可以查询多个提交者邮箱，为空时就是空
# authors = get_authors()
def get_authors():
    authors = input("请输入提交者Emails(多个数据时，请用|分隔): ")

    if len(authors) != 0:
        authors = authors.split('|')
        print(f"您输入的 提交者Emails 是 {authors}")
    else:
        print(f"您输入的 提交者Emails 为空")
    return authors


# 14_CodeIssue_01_code_issue_list
# 获取规则key，一次可以查询多个规则，为空时就是空
# rules = get_rules()
def get_rules():
    rules = input("请输入规则keys(多个数据时，请用|分隔): ")

    if len(rules) != 0:
        rules = rules.split('|')
        print(f"您输入的 规则keys 是 {rules}")
    else:
        print(f"您输入的 规则keys 为空")
    return rules


# 14_CodeIssue_01_code_issue_list
# 获取语言:CloudFormation,CPP,CSS,Go,HTML,Java,JavaScript,Kotlin,PHP,Python,Ruby,Scala,Terraform,TypeScript,XML从上述语言中选择一个或多个，或者不选择就是默认全部，为空时就是空
# languages = get_languages()
def get_languages():
    languages = input("请输入语言s(多个数据时，请用|分隔): ")

    if len(languages) != 0:
        languages = languages.split('|')
        print(f"您输入的 语言s 是 {languages}")
    else:
        print(f"您输入的 语言s 为空")


    # 这里进行了一个循环遍历，将languages中的每一个元素遍历一遍，然后找到language_and_rules中，key是这个元素的所有value值，然后将这些value，组装成一个新的list
    result_list = []
    if len(languages) != 0:
        languages_list = languages
        language_and_rules = get_language_and_rules()
        # print(language_and_rules)
        # 初始化一个空列表，用于存放结果

        # 遍历列表元素
        for element in languages_list:
            # 检查字典中是否有对应的键
            if element in language_and_rules:
                # 将对应键的值添加到结果列表中
                result_list.extend(language_and_rules[element])

    return languages, result_list


# 保存用户输入的排名类型
def get_rankingType():
    rankingType = input("请输入排名类型,支持'test_coverage',\
'comment_coverage','code_quality','dev_equivalent','dev_value','overall',若不输入使用上述六种排名类型全部查询: ")
    while rankingType not in ["test_coverage","comment_coverage","code_quality","dev_equivalent","dev_value","overall"] and len(rankingType) != 0:
        rankingType = input(f"请重新输入排名类型,支持'test_coverage',\
'comment_coverage','code_quality','dev_equivalent','dev_value','overall',若不输入使用上述六种排名类型全部查询:  ")

    if len(rankingType) != 0:
        print(f"您输入的 排名类型 是 {rankingType}")
    else:
        print(f"您输入的 问题类型 为空, 稍后将查询全部六种排名类型的数据")
    return rankingType



# 计算得到最终的规则
# final_rules = get_final_rules(rules,languages)
def get_final_rules(rules,languages):
    if rules == "" and languages == []:
        final_rules = ""
    elif rules != "" and languages != []:
        final_rules = rules + languages
    elif rules != "" and languages == []:
        final_rules = rules
    elif rules == "" and languages != []:
        final_rules = languages
    else:
        pass

    return final_rules


# 获取language_and_rules.md文件的内容
def get_language_and_rules():
    # 打开并读取Markdown文件
    md_file_path = 'language_and_rules.md'

    with open(md_file_path, 'r', encoding='utf-8') as file:
        markdown_content = file.read()

    # 打印Markdown文件的内容
    # print(markdown_content)
    language_and_rules = markdown_content
    language_and_rules = json.loads(language_and_rules)

    return language_and_rules


def get_rules_and_language_key():
    rules_and_language_key = get_rules_and_language()
    # print(rules_and_language_key)
    return rules_and_language_key



# 获取rules_and_language.md文件的内容
def get_rules_and_language():
    # 打开并读取Markdown文件
    md_file_path = 'rules_and_language.md'

    with open(md_file_path, 'r', encoding='utf-8') as file:
        markdown_content = file.read()

    # 打印Markdown文件的内容
    # print(markdown_content)
    rules_and_language = markdown_content
    rules_and_language = json.loads(rules_and_language)

    return rules_and_language



# 获取问题类型，一次可以查询多个问题类型，为空时就是空
# issue_types = get_issue_types()
def get_issue_types():
    issue_types = input("请输入问题类型(多个数据时，请用|分隔),支持'PERFORMANCE',\
'VULNERABILITY','SECURITY_HOTSPOT','PORTABILITY','BUG','CODE_SMELL',不输入时则检索全部问题类型: ")
    while issue_types not in ["PERFORMANCE","VULNERABILITY","SECURITY_HOTSPOT","PORTABILITY","BUG","CODE_SMELL"] and len(issue_types) != 0:
        issue_types = input(f"请重新输入您想检索的查询类型，支持'PERFORMANCE','VULNERABILITY','SECURITY_HOTSPOT','PORTABILITY','BUG','CODE_SMELL',或不输入查询全部问题类型: ")

    if len(issue_types) != 0:
        issue_types = issue_types.split('|')
        print(f"您输入的 问题类型 是 {issue_types}")
    else:
        print(f"您输入的 问题类型 为空")
    return issue_types



# 获取严重程度，一次可以查询多个严重程度，为空时就是空
# severities = get_severities()
def get_severities():
    severities = input("请输入严重程度(多个数据时，请用|分隔),支持'BLOCKER','MINOR','CRITICAL','MAJOR','INFO',不输入时则检索全部严重程度: ")
    while severities not in ["BLOCKER","MINOR","CRITICAL","MAJOR","INFO"] and len(severities) != 0:
        severities = input(f"请重新输入您想检索的查询类型，支持'BLOCKER','MINOR','CRITICAL','MAJOR','INFO',或不输入查询全部严重程度: ")

    if len(severities) != 0:
        severities = severities.split('|')
        print(f"您输入的 严重程度 是 {severities}")
    else:
        print(f"您输入的 严重程度 为空")
    return severities

# 获取文件名（正则表达式匹配），一次只能查询一个文件名，为空时就是空
# filename = get_filename()
def get_filename():
    filename = input("请输入文件名（正则表达式匹配）: ")

    if len(filename) != 0:
        print(f"您输入的 文件名（正则表达式匹配） 是 {filename}")
    else:
        print(f"您输入的 文件名（正则表达式匹配） 为空")
    return filename



# 保存用户输入的一个布尔变量
def get_bool_params(params_name):
    result = input(f"输入是否 {params_name}，可以输入true或者输入false，若不输入则为false: ")
    while len(result) != 0 and not(result in ['true', 'false']):
        result = input(f"请重新输入输入是否 {params_name}，可以输入true或者输入false，若不输入则为false:")
    result = judge_bool(result)

    print(f"您输入的是否 {params_name} 是(空代表false) {result}")
    return result



# 判断一个变量是true 还是false 并返回布尔变量
def judge_bool(string_):
    if isinstance(string_, bool):
        bool_ = string_
    else:
        if string_ not in [None, '', [], ""]:
            string_ = string_.lower()
        if string_ == 'true':
            bool_ = True
        elif string_ == 'false':
            bool_ = False
        else:
            bool_ = False
    # print(f"bool_ = {bool_}")
    return bool_



# for other

# 判断输入内容是否是数字，如果是数字返回True，否则返回False
def is_valid_number(s):
    try:
        float(s)
        return False
    except ValueError:
        pass
    try:
        unicodedata.numeric(s)
        return False
    except (TypeError, ValueError):
        pass
    return True



# 判断输入内容是否是日期格式，如果不是日期格式，返回False，否则返回True
def is_invalid_date(str):
    try:
        datetime.strptime(str, "%Y-%m-%d")
        return False
    except:
        return True



# 将openapi接口导出的str类型的utc时间转换为prc时间字符串
def utc_to_prc_time(utc_time):
    time = datetime.strptime(utc_time,"%Y-%m-%dT%H:%M:%S.000Z")
    prc_time = time + timedelta(hours=8)
    prc_time = prc_time.strftime("%Y-%m-%dT%H:%M:%S.000+08:00")
    return prc_time


# 将prc时间字符串转换为utc时间字符串
def prc_to_utc_time(prc_time):
    time = datetime.strptime(prc_time,"%Y-%m-%d %H:%M:%S")
    utc_time = time - timedelta(hours=8)
    utc_time = utc_time.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    return utc_time


# 将str类型的日期加1天
def add_1_day_to_date(origin_date):
    origin_date = datetime.strptime(origin_date,"%Y-%m-%d")
    final_date = origin_date + timedelta(days=1)
    final_date = final_date.strftime("%Y-%m-%d")
    return final_date


# 判断输入内容是否是日期时间格式，如果是日期时间格式，返回True，否则返回False
def is_invalid_datetime(str):
    try:
        datetime.strptime(str, "%Y-%m-%d %H:%M:%S")
        return False
    except:
        return True



# 当两个字典中的key值相同时，value中dict部分，我想要保留不同的部分
def merge_dicts(dict1, dict2):
    merged_dict = dict1.copy()
    
    for key, value in dict2.items():
        if key in merged_dict and isinstance(value, dict) and isinstance(merged_dict[key], dict):
            merged_dict[key] = merge_dicts(merged_dict[key], value)
        else:
            merged_dict[key] = value
    
    return merged_dict



# for merico_project

# 这个方法是所有项目list格式化为dict
def get_group_list_dict(group_list):
    group_list_dict = {}
    for i in group_list:
        group_list_dict[i['id']] = i

    return group_list_dict




# 这个方法通过递归调用自己，获取指定项目的所有父节点及其父节点的项目ID和项目名称
def get_parent_dict_for_project_id(group_list_dict, project_id, p_dict):
    parent_id = group_list_dict[project_id]['parentId']
    if parent_id:
        # parent_id = parent_id
        # print(f"parent_id = {parent_id}")
        parent_name = group_list_dict[parent_id]['name']
        # print(f"parent_name = {group_list_dict[parent_id]['name']}")
        get_parent_dict_for_project_id(group_list_dict, parent_id, p_dict)
        p_dict[parent_id] = {}
        p_dict[parent_id]['parent_id'] = parent_id
        p_dict[parent_id]['parent_name'] = parent_name
    parent_dict = p_dict.copy()

    # print(f"parent_dict = {parent_dict}\n")
    return parent_dict




# 这个方法给group_list_dict中每个项目dict，增加parent_dict(全部父节点)和children_dict(全部后代子节点)两个key-value的dict
def get_children_dict(group_list_dict):
    for index, i in enumerate(group_list_dict.items()):
        # print(f"i = {i}")
        p_dict = {}
        project_id = i[1]['id']
        parent_dict = get_parent_dict_for_project_id(group_list_dict, project_id, p_dict)
        i[1]['parent_dict'] = parent_dict

    for index_2, i_2 in enumerate(group_list_dict.items()):
        i_2[1]['children_dict'] = {}
        for index_3, i_3 in enumerate(group_list_dict.items()):
            if i_2[1]['id'] in i_3[1]['parent_dict'].keys():
                i_2[1]['children_dict'][i_3[1]['id']] = {}
                i_2[1]['children_dict'][i_3[1]['id']]['id'] = i_3[1]['id']
                i_2[1]['children_dict'][i_3[1]['id']]['name'] = i_3[1]['name']

    # print(f"group_list_dict = {group_list_dict}\n")
    return group_list_dict



# 这个方法是对group_list_dict的parent_dict字段进行一些优化显示
def get_parent_dict_detail(group_list_dict):
    parent_level_count = 0
    for index_6, i_6 in enumerate(group_list_dict.items()):
        i_6[1]['项目层级'] = (len(i_6[1]['parent_dict']) + 1)

    for index_7, i_7 in enumerate(group_list_dict.items()):
        self_project_name = i_7[1]['name']
        project_structure_list = []
        for index_8, i_8 in enumerate(i_7[1]['parent_dict'].items()):
            project_structure_list.append(i_8[1]['parent_name'])
        project_structure_list.append(self_project_name)
        project_structure_list_string = ">".join(map(str, project_structure_list))
        i_7[1]['项目架构'] = project_structure_list_string


    for index, i in enumerate(group_list_dict.items()):
        if len(i[1]['parent_dict']) > parent_level_count:
            parent_level_count = len(i[1]['parent_dict'])
            # print(f"parent_level_count = {parent_level_count}")

    for index_2, i_2 in enumerate(group_list_dict.items()):
        for i_3 in range(parent_level_count):
            i_2[1][f'{i_3+1}级父项目ID'] = None
            i_2[1][f'{i_3+1}级父项目名称'] = None

    for index_4, i_4 in enumerate(group_list_dict.items()):
        for index_5, i_5 in enumerate(i_4[1]['parent_dict'].items()):
            i_4[1][f'{index_5+1}级父项目ID'] = i_5[1]['parent_id']
            i_4[1][f'{index_5+1}级父项目名称'] = i_5[1]['parent_name']



    group_list_dict = {k: v for k, v in sorted(group_list_dict.items(), key=lambda item:(item[1]["项目层级"], item[1]["项目架构"]), reverse=False)}

    # print(f"group_list_dict = {group_list_dict}")
    # print(f"****************************************")
    # print(f"parent_level_count = {parent_level_count}")
    # print(f"恭喜成功导出项目列表: 项目数量 {len(group_list_dict)} ")
    # print(f"================================================")

    return group_list_dict, parent_level_count



# 这个方法是获取指定项目的 直属子节点项目 或者 全部后代节点子项目
def get_project_statistical_dict(special_project_statistical_mode, project_pattern_dict):
    # if special_project_statistical_mode == 'child nodes only':
    #     project_statistical_dict = {}
    #     for index, i in enumerate(project_pattern_dict.items()):

    # elif special_project_statistical_mode == 'all descendant nodes':
    #     project_statistical_dict = {}
    #     for index, i in enumerate(project_pattern_dict.items()):
    #         project_statistical_dict[i[1]['id']] = i[1]
    #         children_dict = i[1]['children_dict']
    #         for index_2, i_2 in enumerate(children_dict.items()):
    #             project_statistical_dict[i_2[1]['id']] = i_2[1]

    project_statistical_dict = {}
    for index, i in enumerate(project_pattern_dict.items()):
        project_statistical_dict[i[1]['id']] = i[1]
        if special_project_statistical_mode == 'all descendant nodes':
            children_dict = i[1]['children_dict']
            for index_2, i_2 in enumerate(children_dict.items()):
                project_statistical_dict[i_2[1]['id']] = i_2[1]
        elif special_project_statistical_mode == 'child nodes only':
            children_dict = i[1]['direct_child']
            for index_2, i_2 in enumerate(children_dict.items()):
                project_statistical_dict[i_2[1]['id']] = i_2[1]

    # print(f"project_statistical_dict = {project_statistical_dict}\n")
    return project_statistical_dict




#这个函数，给某一个project_list 使用，返回每个project的id、name、description、parentId、parent_dict、children_dict、
# direct_child、project_level、project_hierarchy、parent_hierarchy、以及每一级父项目的ID 和 父项目的名称。

from collections import OrderedDict

def build_project_structure(projects):
    # 将输入列表转换为字典，方便按 id 查询
    project_dict = {project['id']: project for project in projects}
    
    # 初始化每个项目的 parent_dict 和 children_dict 以及其他字段
    for project in project_dict.values():
        project['parent_dict'] = {}
        project['children_dict'] = {}
        project['direct_child'] = {}  # 新增字段，存储直属子项目信息
        project['project_level'] = 0
        project['project_hierarchy'] = project['name']
        project['parent_hierarchy'] = []  # 新增字段，用于存储父项目信息
    
    # 更新 parent_dict 和 children_dict
    for project in project_dict.values():
        parent_id = project['parentId']
        if parent_id:
            project['parent_dict'][parent_id] = {
                'parent_id': parent_id,
                'parent_name': project_dict[parent_id]['name']
            }
            project_dict[parent_id]['children_dict'][project['id']] = {
                'id': project['id'],
                'name': project['name']
            }
            # 更新 direct_child 字段
            project_dict[parent_id]['direct_child'][project['id']] = {
                'id': project['id'],
                'name': project['name']
            }
    
    # 初始化最大 parent_hierarchy 长度
    max_parent_hierarchy_length = 0
    
    # 递归更新 project_level, project_hierarchy 和 parent_hierarchy 并累积所有后代节点
    def update_project_info(project_id, level, hierarchy, parent_hierarchy):
        nonlocal max_parent_hierarchy_length
        project = project_dict[project_id]
        project['project_level'] = level
        project['project_hierarchy'] = hierarchy
        project['parent_hierarchy'] = parent_hierarchy.copy()
        
        # 更新最大 parent_hierarchy 长度
        max_parent_hierarchy_length = max(max_parent_hierarchy_length, len(parent_hierarchy))
        
        all_descendants = {}  # 累积所有后代节点
        
        for child in project['children_dict'].values():
            child_id = child['id']
            new_parent_hierarchy = parent_hierarchy + [{'parent_id': project['id'], 'parent_name': project['name']}]
            update_project_info(child_id, level + 1, hierarchy + ' > ' + project_dict[child_id]['name'], new_parent_hierarchy)
            
            # 将子节点及其所有后代添加到 all_descendants
            all_descendants[child_id] = child
            all_descendants.update(project_dict[child_id]['children_dict'])
        
        project['children_dict'] = all_descendants
    
    # 从根节点开始更新
    for project in project_dict.values():
        if not project['parentId']:
            update_project_info(project['id'], 1, project['name'], [])
    
    # 添加父项目信息字段
    for project in project_dict.values():
        for i in range(max_parent_hierarchy_length):
            parent_level = i + 1
            if i < len(project['parent_hierarchy']):
                parent_info = project['parent_hierarchy'][i]
                project[f'{parent_level}级父项目ID'] = parent_info['parent_id']
                project[f'{parent_level}级父项目名称'] = parent_info['parent_name']
            else:
                project[f'{parent_level}级父项目ID'] = ''
                project[f'{parent_level}级父项目名称'] = ''
    
    # 按 project_level 和 project_hierarchy 进行排序
    sorted_project_items = sorted(project_dict.items(), key=lambda x: (x[1]['project_level'], x[1]['project_hierarchy']))
    
    # 构建有序字典
    sorted_project_dict = OrderedDict(sorted_project_items)
    
    return sorted_project_dict, max_parent_hierarchy_length

# # 示例使用
# projects = [
#     {"id": "9dd149f1-475b-44ae-89a6-acac807f8cc6", "name": "Merico", "description": "", "parentId": None},
#     {"id": "c6be658c-210d-4a30-a2fa-16311646d2bb", "name": "Tech", "description": None, "parentId": "9dd149f1-475b-44ae-89a6-acac807f8cc6"},
#     {"id": "987bb56c-0ff2-4a78-a6e5-7ceaed961041", "name": "Sales", "description": None, "parentId": "9dd149f1-475b-44ae-89a6-acac807f8cc6"},
#     {"id": "36967f48-91f8-49ab-a4fa-d459e18d8d08", "name": "EE Code", "description": "", "parentId": "8388b6aa-658f-4c7b-9a68-b3ae724f75b8"},
#     {"id": "a8b76e43-0c38-4cb0-b48d-159bde92e285", "name": "AE Code", "description": "", "parentId": "8388b6aa-658f-4c7b-9a68-b3ae724f75b8"},
#     {"id": "8388b6aa-658f-4c7b-9a68-b3ae724f75b8", "name": "Merico-3-level", "description": "", "parentId": "c6be658c-210d-4a30-a2fa-16311646d2bb"},
#     {"id": "336380b1-f8ca-4811-aadb-d9bfb270c8c3", "name": "Service", "description": None, "parentId": "9dd149f1-475b-44ae-89a6-acac807f8cc6"},
#     {"id": "1fe1d6b0-502b-4585-87f2-f833648c519f", "name": "Finance", "description": None, "parentId": "9dd149f1-475b-44ae-89a6-acac807f8cc6"}
# ]

# sorted_project_dict, max_parent_hierarchy_length = build_project_structure(projects)

# # 打印结果
# for project_id, project_info in sorted_project_dict.items():
#     print(f"ID: {project_info['id']}")
#     print(f"Name: {project_info['name']}")
#     print(f"Description: {project_info['description']}")
#     print(f"Parent ID: {project_info['parentId']}")
#     print(f"Project Level: {project_info['project_level']}")
#     print(f"Project Hierarchy: {project_info['project_hierarchy']}")
#     print(f"Parent Dict: {project_info['parent_dict']}")
#     print(f"Children Dict: {project_info['children_dict']}")
#     print(f"Direct Child: {project_info['direct_child']}")
#     print(f"Parent Hierarchy: {project_info['parent_hierarchy']}")
#     for i in range(max_parent_hierarchy_length):
#         parent_level = i + 1
#         print(f"{parent_level}级父项目ID: {project_info[f'{parent_level}级父项目ID']}")
#         print(f"{parent_level}级父项目名称: {project_info[f'{parent_level}级父项目名称']}")
#     print()

# print(f"Max Parent Hierarchy Length: {max_parent_hierarchy_length}")




# for merico repo
# 这个方法是所有代码库list格式化为dict
def get_repo_list_dict(repo_list):
    repo_list_dict = {}
    for i in repo_list:
        repo_list_dict[i['id']] = i

    return repo_list_dict



# for merico_team

# 这个方法是所有项目list格式化为dict
def get_team_list_dict(team_list):
    team_list_dict = {}
    for i in team_list:
        team_list_dict[i['id']] = i

    return team_list_dict



# 有team_list，这个list，描述的是一个树状结构的team 信息，返回每一个team 的id、name、parentId、parent_dict、children_dict、
# team层级(就是这个team在第几级)、team架构(就是这个team从顶级开始直到自己的层级关系用>串联)

# children_dict，返回team 的全部后代节点的team信息，把这个team全部后代节点的team都返回出来，具体来说就是包括儿子team、孙子team 以及后续的全部后代节点team

# 在这个函数build_team_structure的返回中，给每一个team 假设这个team叫做team_A，再新增一个direct_child的字段，这个字段使用dict方式保存数据，保存的是直属子团队的数据

# 这个dict中的team，都有一个特点，那就是他们每一个team的 parentId，恰好就是team_A 的id

# dict中的每一个元素，key是team的id，value也是一个dict，dict使用类似这种方式保存：{'id': 'd483b8ef-44b5-45c1-bb50-2420fe0667c6', 'name': 'E1-1-1'}
from collections import OrderedDict

def build_team_structure(teams):
    # 将输入列表转换为字典，方便按 id 查询
    team_dict = {team['id']: team for team in teams}
    
    # 初始化每个团队的 parent_dict 和 children_dict 以及其他字段
    for team in team_dict.values():
        team['parent_dict'] = {}
        team['children_dict'] = {}
        team['direct_child'] = {}  # 新增字段，存储直属子团队信息
        team['team_level'] = 0
        team['team_hierarchy'] = team['name']
        team['parent_hierarchy'] = []  # 新增字段，用于存储父团队信息
    
    # 更新 parent_dict 和 children_dict
    for team in team_dict.values():
        parent_id = team['parentId']
        if parent_id:
            team['parent_dict'][parent_id] = {
                'parent_id': parent_id,
                'parent_name': team_dict[parent_id]['name']
            }
            team_dict[parent_id]['children_dict'][team['id']] = {
                'id': team['id'],
                'name': team['name']
            }
            # 更新 direct_child 字段
            team_dict[parent_id]['direct_child'][team['id']] = {
                'id': team['id'],
                'name': team['name']
            }
    
    # 初始化最大 parent_hierarchy 长度
    max_parent_hierarchy_length = 0
    
    # 递归更新 team_level, team_hierarchy 和 parent_hierarchy 并累积所有后代节点
    def update_team_info(team_id, level, hierarchy, parent_hierarchy):
        nonlocal max_parent_hierarchy_length
        team = team_dict[team_id]
        team['team_level'] = level
        team['team_hierarchy'] = hierarchy
        team['parent_hierarchy'] = parent_hierarchy.copy()
        
        # 更新最大 parent_hierarchy 长度
        max_parent_hierarchy_length = max(max_parent_hierarchy_length, len(parent_hierarchy))
        
        all_descendants = {}  # 累积所有后代节点
        
        for child in team['children_dict'].values():
            child_id = child['id']
            new_parent_hierarchy = parent_hierarchy + [{'parent_id': team['id'], 'parent_name': team['name']}]
            update_team_info(child_id, level + 1, hierarchy + ' > ' + team_dict[child_id]['name'], new_parent_hierarchy)
            
            # 将子节点及其所有后代添加到 all_descendants
            all_descendants[child_id] = child
            all_descendants.update(team_dict[child_id]['children_dict'])
        
        team['children_dict'] = all_descendants
    
    # 从根节点开始更新
    for team in team_dict.values():
        if not team['parentId']:
            update_team_info(team['id'], 1, team['name'], [])
    
    # 添加父团队信息字段
    for team in team_dict.values():
        for i in range(max_parent_hierarchy_length):
            parent_level = i + 1
            if i < len(team['parent_hierarchy']):
                parent_info = team['parent_hierarchy'][i]
                team[f'{parent_level}级父团队ID'] = parent_info['parent_id']
                team[f'{parent_level}级父团队名称'] = parent_info['parent_name']
            else:
                team[f'{parent_level}级父团队ID'] = ''
                team[f'{parent_level}级父团队名称'] = ''
    
    # 按 team_level 和 team_hierarchy 进行排序
    sorted_team_items = sorted(team_dict.items(), key=lambda x: (x[1]['team_level'], x[1]['team_hierarchy']))
    
    # 构建有序字典
    sorted_team_dict = OrderedDict(sorted_team_items)
    
    return sorted_team_dict, max_parent_hierarchy_length

# # 示例使用
# teams = [
#     {"id": "d9526a7a-2f84-4ddf-a3f9-f74b8e8c455d", "name": "Merico", "parentId": None},
#     {"id": "0cfb6cb5-9458-4937-88eb-5b26fd712cac", "name": "Engineer", "parentId": "d9526a7a-2f84-4ddf-a3f9-f74b8e8c455d"},
#     {"id": "0e6f42f3-c918-4c20-b909-56f7e836ca61", "name": "Service", "parentId": "d9526a7a-2f84-4ddf-a3f9-f74b8e8c455d"},
#     {"id": "cc422dc9-fd08-4707-b6f0-c879bc3d7f0d", "name": "E1", "parentId": "0cfb6cb5-9458-4937-88eb-5b26fd712cac"},
#     {"id": "947bdc3e-5b17-456f-b15b-9eb4eae4e300", "name": "E1-1", "parentId": "cc422dc9-fd08-4707-b6f0-c879bc3d7f0d"},
#     {"id": "928253a1-3042-4e0e-b50a-9c3ab263b789", "name": "S-1", "parentId": "0e6f42f3-c918-4c20-b909-56f7e836ca61"},
#     {"id": "36b4a10e-76a3-41a8-a627-df11c828282d", "name": "S-2", "parentId": "0e6f42f3-c918-4c20-b909-56f7e836ca61"},
#     {"id": "d483b8ef-44b5-45c1-bb50-2420fe0667c6", "name": "E1-1-1", "parentId": "947bdc3e-5b17-456f-b15b-9eb4eae4e300"}
# ]

# sorted_team_dict, max_parent_hierarchy_length = build_team_structure(teams)

# # 打印结果
# for team_id, team_info in sorted_team_dict.items():
#     print(f"ID: {team_info['id']}")
#     print(f"Name: {team_info['name']}")
#     print(f"Parent ID: {team_info['parentId']}")
#     print(f"Team Level: {team_info['team_level']}")
#     print(f"Team Hierarchy: {team_info['team_hierarchy']}")
#     print(f"Parent Dict: {team_info['parent_dict']}")
#     print(f"Children Dict: {team_info['children_dict']}")
#     print(f"Direct Child: {team_info['direct_child']}")
#     print(f"Parent Hierarchy: {team_info['parent_hierarchy']}")
#     for i in range(max_parent_hierarchy_length):
#         parent_level = i + 1
#         print(f"{parent_level}级父团队ID: {team_info[f'{parent_level}级父团队ID']}")
#         print(f"{parent_level}级父团队名称: {team_info[f'{parent_level}级父团队名称']}")
#     print()

# print(f"Max Parent Hierarchy Length: {max_parent_hierarchy_length}")




# 保存输入匹配正则的团队ID或团队名称
def get_team_id_or_name_pattern():
    # 以下内容根据需要填写
    special_team_regex = input("请输入特定的团队ID 或 团队名称(支持正则匹配, 间隔符号仅支持英文|) 或 all(代表查询全部团队), 不支持输入为空: ")
    while len(special_team_regex) == 0:
        special_team_regex = input("请重新输入特定的团队ID 或 团队名称(支持正则匹配, 间隔符号仅支持英文|) 或 all(代表查询全部团队), 不支持输入为空: ")

    team_pattern = re.compile(special_team_regex, re.IGNORECASE )

    if len(special_team_regex) != 'all':
        print(f"您输入的 特定的团队(组)ID 或 团队名称 是 {special_team_regex}")
    else:
        print(f"您输入的是 'all', 稍后将查询全部团队(组)数据")

    return special_team_regex, team_pattern



# 保存用户输入的查询团队成员列表时，是否包含子团队
def get_sub_team():
    result = input(f"查询团队成员列表时，是否包含子团队，可以输入true或者输入false，若不输入则为false: ")
    while len(result) != 0 and not(result in ['true', 'false']):
        result = input(f"请重新输入查询团队成员列表时，是否包含子团队，可以输入true或者输入false，若不输入则为false:")

    if result == 'true':
        result = True
    elif result == 'false' or len(result) == 0:
        result = False

    print(f"查询团队成员列表时，是否包含子团队 是 {result}")
    return result



# 使用团队ID 或 团队名称 或 all 检索团队
def find_team_by_id_or_name(special_team_regex, team_pattern, group_list_dict):
    # 要查询的团队的ids
    # team_list = get_team_list(client)
    # print(f"group_list_dict = {group_list_dict}")
    # list to dict
    team_dict = {}

    if special_team_regex == 'all':
        for index, i in enumerate(group_list_dict.items()):
            team_dict[i[1]['id']] = {}
            team_dict[i[1]['id']] = i[1]
            team_dict[i[1]['id']]['团队ID'] = i[1]['id']
            team_dict[i[1]['id']]['检索命中字段'] = "列出全部团队"
            team_dict[i[1]['id']]['被检索字段'] = "列出全部团队"
            team_dict[i[1]['id']]['团队名称'] = i[1]['name']

    else:
        for index_2, i_2 in enumerate(group_list_dict.items()):
            if team_pattern.findall(i_2[1]['id']):
                team_dict[i_2[1]['id']] = {}
                team_dict[i_2[1]['id']] = i_2[1]
                team_dict[i_2[1]['id']]['团队ID'] = i_2[1]['id']
                team_dict[i_2[1]['id']]['检索命中字段'] = '团队ID'

            if team_pattern.findall(i_2[1]['name']):
                if i_2[1]['id'] in team_dict.keys():
                    origin_str = team_dict[i_2[1]['id']]['检索命中字段']
                    team_dict[i_2[1]['id']]['检索命中字段'] = f"{origin_str} & 团队名称"
                else:
                    team_dict[i_2[1]['id']] = {}
                    team_dict[i_2[1]['id']] = i_2[1]
                    team_dict[i_2[1]['id']]['团队ID'] = i_2[1]['id']
                    team_dict[i_2[1]['id']]['检索命中字段'] = '团队名称'

            if team_pattern.findall(i_2[1]['id']) or team_pattern.findall(i_2[1]['name']):
                team_dict[i_2[1]['id']]['被检索字段'] = special_team_regex
                team_dict[i_2[1]['id']]['团队名称'] = i_2[1]['name']



    if len(team_dict) == 0:
        print(f"!!!!!!很抱歉, 系统没有找到匹配团队ID、团队名称 正则表达式的: {special_team_regex} 的团队, 请您重新执行脚本文件, 并修改团队信息, 以便找到您需要的团队")

    print(f"目前累计检索到的团队数量是: {len(team_dict.keys())}")
    # print(f"/n")
    # print(f"team_dict = {team_dict}")
    return team_dict





# for xlsxwriter

# 这个方法主要是给worksheet中写入数据用的
def write_worksheet(worksheet, data):
    # 表头
    row = 0
    col = 0
    for index, i in enumerate(data.items()):
        if index == 0:
            # print(f"i[1] = {i[1]}")
            single_row = i[1].keys()
            single_row = list(single_row)
            length = len(single_row)
            for i_2 in range(length):
                worksheet.write (row, col+i_2, single_row[i_2])
            row += 1

    # 数据
    row = 1
    col = 0
    for index, i in enumerate(data.items()):
        # print(f"i[1] = {i[1]}")
        single_row = i[1].values()
        single_row = list(single_row)
        length = len(single_row)
        for i_2 in range(length):
            # print(f"single_row[i_2] = {single_row[i_2]}")
            try:
                worksheet.write (row, col+i_2, single_row[i_2])
            except TypeError as e:
                # print(e)
                worksheet.write (row, col+i_2, str(single_row[i_2]))
        row += 1

    # print(row)



# for gitlab_api
# https://docs.gitlab.com/ee/api/api_resources.html

def gitlab_request(endpoint, path, params, headers):
    url = f'{endpoint}/{path.lstrip("/")}'
    # print(f"gitlab_request_url = {url}")
    # print(f"gitlab_request_params = {params}")
    # print(f"gitlab_request_headers = {headers}")
    response = requests.get(url, params=params, headers=headers)
    if response.status_code != 200:
        raise Exception(f"request got a error: {response}")
    result = response.json()
    return result



# for posgres

def config_database(host, port, user, password, database):
    # print(f'开始配置： config_database_info')
    # 获得连接
    conn = psycopg2.connect(database=database, user=user, password=password, host=host, port=port)
    # 获得游标对象，一个游标对象可以对数据库进行执行操作
    cursor = conn.cursor()
    # print(f'配置成功！ config_database_info')
    print(f'数据库配置中，请稍等...\n')
    return cursor



# for read excel && csv
# def read_file(file_path):
#     """
#     读取 CSV/Excel 文件并返回行数据，跳过第一行表头
#     """
#     try:
#         rows = []
#         if file_path.endswith(".csv"):
#             with open(file_path, mode="r", encoding="utf-8") as f:
#                 reader = csv.reader(f)
#                 next(reader, None)  # 跳过第一行表头
#                 rows = [row for row in reader]
#         elif file_path.endswith(".xlsx"):
#             workbook = openpyxl.load_workbook(file_path)
#             sheet = workbook.active
#             rows = [[cell.value for cell in row] for row in sheet.iter_rows()][1:]  # 跳过第一行表头
#         else:
#             raise ValueError("Unsupported file format.")
#         return rows
#     except Exception as e:
#         logging.error(f"读取文件出错: {e}")
#         sys.exit(1)

# for read excel && csv
def read_file(file_path):
    """
    读取 CSV/Excel 文件并返回行数据，跳过第一行表头
    支持自动检测CSV文件编码
    """
    try:
        rows = []
        if file_path.endswith(".csv"):
            # 先检测文件编码
            with open(file_path, 'rb') as f:
                raw_data = f.read()
                encoding = detect(raw_data)['encoding']
            # 尝试用检测到的编码打开文件
            try:
                with open(file_path, mode="r", encoding=encoding) as f:
                    reader = csv.reader(f)
                    next(reader, None)  # 跳过第一行表头
                    rows = [row for row in reader]
            except UnicodeDecodeError:
                # 如果检测的编码失败，尝试常见的中文编码
                encodings = ['gbk', 'gb2312', 'utf-16', 'utf-8']
                for enc in encodings:
                    try:
                        with open(file_path, mode="r", encoding=enc) as f:
                            reader = csv.reader(f)
                            next(reader, None)
                            rows = [row for row in reader]
                            break
                    except UnicodeDecodeError:
                        continue
                else:
                    raise ValueError(f"无法解码CSV文件，尝试的编码: {encoding}, {', '.join(encodings)}")
        elif file_path.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            rows = [[cell.value for cell in row] for row in sheet.iter_rows()][1:]  # 跳过第一行表头
        else:
            raise ValueError("不支持的文件格式，仅支持 .csv 和 .xlsx 文件")
        return rows
    except Exception as e:
        logging.error(f"读取文件 {file_path} 出错: {str(e)}")
        sys.exit(1)


# 读取csv文件中的内容，并且生成一个dict，dict中使用二级嵌套方式保存数据，第一级的key，就使用csv 的第一列的内容，value也是dict
# 第二级的每一个key-value，分别对应csv中的表头 和从第二行开始的数据的每一个数据
def read_file_to_dict(file_path):
    # 检查文件路径是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")
    
    _, file_extension = os.path.splitext(file_path)
    
    # 读取数据
    if file_extension == '.csv':
        with open(file_path, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            data = [row for row in reader]
    elif file_extension in ['.xls', '.xlsx']:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows())]
        data = [
            {headers[i]: cell.value for i, cell in enumerate(row)}
            for row in sheet.iter_rows(min_row=2)
        ]
    else:
        raise ValueError("Unsupported file format")
    
    # 构造字典
    data_dict = {}
    for row in data:
        first_column = row.get('key')
        if first_column:
            if first_column not in data_dict:
                data_dict[first_column] = {}
            for key, value in row.items():
                if key != 'key':
                    data_dict[first_column][key] = value
    return data_dict



# for timestamp
# 这个函数的作用是将几种形式的时间表达参数，都转换成string类型的timestamp(毫秒级)数据
# 转换为毫秒级时间戳
def convert_to_timestamp(time_str):
    try:
        dt = datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
        timestamp_ms = int(dt.timestamp() * 1000)
        return timestamp_ms
    except ValueError:
        return None

# 这个函数的作用是将几种形式的时间表达参数，都转换成string类型的timestamp(秒级)数据
# 转换为秒级时间戳
def convert_time_to_timestamp_with_second(time_str):
    try:
        dt = datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
        timestamp_s = int(dt.timestamp())
        return timestamp_s
    except ValueError:
        return None



# for string
# 这个函数的作用是将自定义标签中{"name":"员工类型","value":"Bosch"}，转换成[{"name":"员工类型","value":"Bosch"}]
def convert_json_string_to_list(json_string):
    try:
        # 如果输入字符串为空，则直接返回空列表
        if not json_string:
            return None

        # 使用 json.loads() 函数将 JSON 字符串解析为 Python 对象
        json_list = json.loads('[' + json_string + ']')
        return json_list
    except json.JSONDecodeError:
        # 解析失败时返回空列表
        return []



# 这个函数的作用是将用户副邮箱地址中的'chao.cheng_up@meri.co,chao.cheng_up_2@meri.co'，转换成['chao.cheng_up@meri.co','chao.cheng_up_2@meri.co']
def convert_string_to_list(string):
    # 如果输入字符串为空或者为None，则直接返回空列表
    if string is None or not string.strip():
        return []
    
    # 使用 split() 函数将字符串按逗号分割成子字符串，并去除空格
    result_list = [s.strip() for s in string.split(',')]
    return result_list



# 将list 转换成dict，并且使用 Python 代码将列表转换为字典，其中每个元素的 id 和 一个 20个字符的随机字符串 作为键
def convert_list_to_dict(data_list, key):
    data_dict = {}
    for item in data_list:
        ran_str = ''.join(random.sample(string.ascii_letters + string.digits, 20))
        data_dict[f"{item[key]}-{ran_str}"] = item
    return data_dict



# 将list 转换成dict，并且使用 Python 代码将列表转换为字典，其中每个元素的 id 作为键
def convert_list_to_dict_without_random_str(data_list, key):
    data_dict = {}
    for item in data_list:
        data_dict[f"{item[key]}"] = item
    return data_dict



# 将string 类型字符串转换成number
def convert_str_to_number(string):
    number = ''
    if string is None or not string.strip():
        pass
    else:
        try:
            number = int(string)
        except Exception as e:
            print(f"Error convert_str_to_number: {e}")
            pass
    return number



# 使用 Python 代码来遍历 target_dict 并检查每个键是否存在于 source_dict 中。如果存在，将 source_dict 中对应的值赋给 target_dict
def update_dict_from_another(target_dict, source_dict):
    """
    将 target_dict 中存在于 source_dict 的键的值更新为 source_dict 中对应的值

    参数:
    target_dict (dict): 需要更新的目标字典
    source_dict (dict): 用于更新的源字典

    返回:
    dict: 更新后的目标字典
    """
    # 遍历 target_dict 的键
    for key in target_dict.keys():
        # 检查该键是否存在于 source_dict 中
        if key in source_dict:
            # 将 source_dict 中对应的值赋给 target_dict
            target_dict[key] = source_dict[key]
    
    return target_dict



# 有两个二级嵌套dict，两个dict的二级嵌套中的dict中有一些相同的key，但是也有不同的key，把这两个dict 进行合并，新合并的dict，同时包含两个dict中相同的key以及不同的key, 已知当两个dict 的二级嵌套dict中 使用相同的key时，对应的value是一样的
# 如果 `dict_1` 和 `dict_2` 的一级嵌套的键值对数量不一致，并且你想要将它们合并成一个新的字典 `dict_3`，那么最终得到的 `dict_3` 的长度将等于 `dict_1` 和 `dict_2` 中键的并集的长度。
# 具体来说：
# - 如果 `dict_1` 中有一些键，在 `dict_2` 中不存在，那么这些键会全部加入到 `dict_3` 中。
# - 如果 `dict_2` 中有一些键，在 `dict_1` 中不存在，那么同样会将这些键加入到 `dict_3` 中。
# 因此，最终 `dict_3` 的长度将等于 `len(dict_1)` 和 `len(dict_2)` 中的较大值，因为它包含了两个字典的所有键。即使两者中某个字典的键比另一个多，最终合并后的字典 `dict_3` 会包含所有这些键。
# 例如，如果 `dict_1` 中有 5 个键，而 `dict_2` 中有 3 个键，合并后的 `dict_3` 将会有 5 个键。
def merge_dict_depend_longer_len(dict_1, dict_2):
    # 初始化一个空的 dict_3 来存放合并结果
    dict_3 = {}

    # 遍历 dict_1 并将其添加到 dict_3 中
    for outer_key, outer_value in dict_1.items():
        dict_3[outer_key] = outer_value

    # 遍历 dict_2 并将其添加到 dict_3 中
    for outer_key, outer_value in dict_2.items():
        if outer_key not in dict_3:
            dict_3[outer_key] = outer_value
        else:
            # 合并内部字典
            for inner_key, inner_value in outer_value.items():
                dict_3[outer_key][inner_key] = inner_value

    return dict_3



# 构造一个函数，使得合并后的 dict_3 中只包含 dict_1 和 dict_2 中相同的一级嵌套键值对，并且只保留两者长度较小的部分
def merge_dict_depend_shorter_len(dict_1, dict_2):
    # 确定较小长度的字典
    min_length = min(len(dict_1), len(dict_2))
    
    # 初始化空字典用于存放合并结果
    dict_3 = {}
    
    # 遍历较小长度的字典，将相同的一级嵌套键值对合并到 dict_3 中
    for key in list(dict_1.keys())[:min_length]:
        if key in dict_2:
            dict_3[key] = dict_1[key]
            dict_3[key].update(dict_2[key])  # 合并相同 key 的值
    
    return dict_3



# 把dict中所有的二级嵌套中的key都保持一致，也就是说如果二级嵌套的dict中缺失key，就添加上这个key，并且对应value使用''
def ensure_keys_in_nested_dicts(data_dict):
    """
    确保嵌套字典中的所有二级嵌套字典包含相同的键，并按照指定顺序排列。
    
    参数:
    data_dict (dict): 包含嵌套字典的字典
    
    返回:
    dict: 处理后的字典，其中所有二级嵌套字典包含相同的键并按指定顺序排列
    """
    # 动态获取所有二级嵌套字典中的键
    required_keys = set()
    for value in data_dict.values():
        required_keys.update(value.keys())
    
    # 将required_keys转为列表并按名称排序，以便后续统一顺序
    required_keys = sorted(list(required_keys))

    # 遍历每个主键的值
    for key, value in data_dict.items():
        # 确保所有key都存在
        for required_key in required_keys:
            if required_key not in value:
                value[required_key] = ''  # 使用空字符串作为默认值
        # 按照required_keys顺序重新排列
        data_dict[key] = OrderedDict((k, value[k]) for k in required_keys)
    
    return data_dict


# 把dict中所有的二级嵌套中的key都保持一致，也就是说查询所有二级嵌套的dict中，共同使用的key有哪些，如果某一个dict中有其他dict没有的key，就删除这个key
def unify_keys_in_nested_dicts(data_dict):
    """
    确保嵌套字典中的所有二级嵌套字典只包含共同的键。
    
    参数:
    data_dict (dict): 包含嵌套字典的字典
    
    返回:
    dict: 处理后的字典，其中所有二级嵌套字典只包含共同的键
    """
    # 找到所有二级嵌套字典的共同键
    common_keys = set(data_dict[next(iter(data_dict))].keys())
    for value in data_dict.values():
        common_keys.intersection_update(value.keys())
    
    # 保留每个二级嵌套字典中的共同键，删除其他键
    for key, value in data_dict.items():
        for k in list(value.keys()):
            if k not in common_keys:
                del value[k]
    
    return data_dict


